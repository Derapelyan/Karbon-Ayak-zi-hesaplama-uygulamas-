"""
UNSPED Carbon Footprint Importer v4
=====================================
- Excel sadece veri girişi — tüm hesaplar burada yapılır
- Birim dönüşümü tam dinamik: m³/L/kg/ton, kWh/MWh/GWh, km/mil, USD/EUR/GBP/TL
- Her satır kendi yılına kaydedilir
- Aynı yıl tekrar import edilirse silinip yeniden yazılır
- Legend/not satırları otomatik atlanır

Kullanım:
    python pipeline/Importer.py data/UNSPED_Karbon_Veri_Girisi_v3.xlsx
"""

import sys, os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from collections import defaultdict

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from db.connection import SessionLocal
from db.models import (
    Company, ReportingPeriod, AuditLog,
    StationaryCombustion, MobileCombustion, Refrigerant,
    ElectricityConsumption, FreightEmission, EmployeeCommuting,
    PurchasedGoods, CapitalGoods
)


# ══════════════════════════════════════════════════════════════════
# BİRİM DÖNÜŞÜM SİSTEMİ
# ══════════════════════════════════════════════════════════════════

conversion_warnings = []

# Kütleyi TON'a çeviren katsayılar
TO_TON = {
    "ton": 1.0, "tonne": 1.0, "t": 1.0,
    "kg":  0.001,
    "g":   0.000001,
    "lb":  0.000453592,
    "gg":  1000.0,
}

# Hacmi LİTRE'ye çeviren katsayılar
TO_LITRE = {
    "l": 1.0, "lt": 1.0, "litre": 1.0, "liter": 1.0,
    "ml": 0.001,
    # m³ ve m3 buradan kaldırıldı — calc_combustion'da ayrı dal ile handle edilir
    # m³ → density (ton/m³) ile direkt tona çevrilir, litre'ye çevrilmez
}

# Enerjiyi MWh'e çeviren katsayılar
TO_MWH = {
    "kwh": 0.001,
    "mwh": 1.0,
    "gwh": 1000.0,
    "twh": 1000000.0,
    "gj":  0.27778,
    "tj":  277.78,
}

# Mesafeyi KM'ye çeviren katsayılar
TO_KM = {
    "km":    1.0,
    "m":     0.001,
    "mil":   1.60934, "mile": 1.60934, "miles": 1.60934,
    "nm":    1.852,
}

# EUR/GBP → USD yaklaşık (sabit kur — TL için Dolar Kuru sütunu kullanılır)
FX_TO_USD = {
    "eur": 1.08, "€": 1.08,
    "gbp": 1.27, "£": 1.27,
    "usd": 1.0,  "$": 1.0,
}

def unit_key(u):
    return str(u or "").lower().strip()

def to_ton(val, unit, label=""):
    """Kütleyi ton'a çevirir. Density gerektiren m³/L için None döner."""
    uk = unit_key(unit)
    if uk in TO_TON:
        factor = TO_TON[uk]
        if factor != 1.0:
            _warn(f"{label}: {val} {unit} → {val*factor:.6f} ton")
        return val * factor
    return None  # m³/L → density gerekiyor

def to_litre(val, unit, label=""):
    uk = unit_key(unit)
    if uk in TO_LITRE:
        factor = TO_LITRE[uk]
        if factor != 1.0:
            _warn(f"{label}: {val} {unit} → {val*factor:.4f} L")
        return val * factor
    return None

def to_mwh(val, unit, label=""):
    uk = unit_key(unit)
    if uk in TO_MWH:
        factor = TO_MWH[uk]
        if factor != 1.0:
            _warn(f"{label}: {val} {unit} → {val*factor:.4f} MWh")
        return val * factor
    return None

def to_km(val, unit, label=""):
    uk = unit_key(unit)
    if uk in TO_KM:
        factor = TO_KM[uk]
        if factor != 1.0:
            _warn(f"{label}: {val} {unit} → {val*factor:.4f} km")
        return val * factor
    return None

def to_usd(val, unit, kur=None, label=""):
    uk = unit_key(unit)
    # USD ve varyantları
    if uk in FX_TO_USD:
        factor = FX_TO_USD[uk]
        if factor != 1.0:
            _warn(f"{label}: {val} {unit} → {val*factor:.4f} USD (~sabit kur)")
        return val * factor
    # TL → USD
    if uk in ("tl", "try", "₺"):
        if kur and kur > 0:
            usd = val / kur
            _warn(f"{label}: {val} TL ÷ {kur} = {usd:.4f} USD")
            return usd
        else:
            print(f"    ⚠ {label}: TL için Dolar Kuru girilmemiş!")
            return 0.0
    # "2021 USD", "2022 USD" gibi yıllı USD
    if "usd" in uk:
        return val
    return val  # bilinmeyen → olduğu gibi

def _warn(msg):
    conversion_warnings.append(msg)
    print(f"    🔄 {msg}")


# ══════════════════════════════════════════════════════════════════
# HESAP MOTORLARI
# ══════════════════════════════════════════════════════════════════

def calc_combustion(amount, unit, ncv, ef_co2, ef_ch4, ef_n2o,
                    gwp_ch4, gwp_n2o, density, label=""):
    """
    Yanma hesabı — tüm birimler desteklenir.
    Adımlar:
      1. Miktarı tona çevir
      2. Ton → Gg (/1000)
      3. Gg × NCV → TJ
      4. TJ × EF → ton CO2e
    """
    try:
        amount  = float(amount  or 0)
        ncv     = float(ncv     or 0)
        ef_co2  = float(ef_co2  or 0)
        ef_ch4  = float(ef_ch4  or 0)
        ef_n2o  = float(ef_n2o  or 0)
        gwp_ch4 = float(gwp_ch4 or 0)
        gwp_n2o = float(gwp_n2o or 0)
        density = float(density or 0)

        if amount == 0 or ncv == 0: return 0.0

        uk = unit_key(unit)

        # Tona çevir — m³ önce (TO_LITRE'de yok)
        if uk in TO_TON:
            amount_ton = to_ton(amount, unit, label)
        elif uk in ("m³", "m3"):
            # m³ direkt density ile: 11064 m³ × 0.000717 ton/m³ = 7.93 ton
            if density == 0:
                print(f"    ⚠ {label}: Yoğunluk girilmemiş!"); return 0.0
            amount_ton = amount * density
        elif uk in TO_LITRE:
            # L → ton via density (ton/L)
            amount_l   = to_litre(amount, unit, label)
            if density == 0:
                print(f"    ⚠ {label}: Yoğunluk girilmemiş!"); return 0.0
            amount_ton = amount_l * density
        else:
            # Bilinmeyen — density ile dene
            if density > 0:
                amount_ton = amount * density
                _warn(f"{label}: bilinmeyen birim '{unit}', density ile çevrildi")
            else:
                print(f"    ⚠ {label}: bilinmeyen birim '{unit}'"); return 0.0

        activity_gg = amount_ton / 1000
        activity_tj = activity_gg * ncv
        co2e = (activity_tj * ef_co2 +
                activity_tj * ef_ch4 * gwp_ch4 +
                activity_tj * ef_n2o * gwp_n2o)
        return round(co2e, 6)
    except Exception as e:
        print(f"    ⚠ Yanma hesap hatası [{label}]: {e}")
        return 0.0

def calc_refrigerant(fv, fv_unit, ef_pct, gwp, label=""):
    """
    Soğutucu gaz: FV × EF × GWP = tCO2e
    FV birimi kg veya ton:
      kg  → FV(kg) × EF × GWP / 1000 = tCO2e
      ton → FV(ton) × EF × GWP       = tCO2e  (ton × GWP = tCO2e direkt)
    EF ondalık girilir: %10 = 0.10
    """
    try:
        fv     = float(fv     or 0)
        ef_pct = float(ef_pct or 0)
        gwp    = float(gwp    or 0)
        if fv == 0 or gwp == 0: return 0.0

        uk = unit_key(fv_unit)
        if uk == "kg":
            emission_kg = fv * ef_pct        # kg gas
            co2e = emission_kg * gwp / 1000  # tCO2e
        elif uk in ("ton", "tonne", "t"):
            emission_ton = fv * ef_pct       # ton gas
            co2e = emission_ton * gwp        # tCO2e
        elif uk == "g":
            emission_g = fv * ef_pct
            co2e = emission_g * gwp / 1_000_000
        else:
            # Bilinmeyen → ton kabul
            _warn(f"{label}: bilinmeyen FV birimi '{fv_unit}', ton kabul edildi")
            emission_ton = fv * ef_pct
            co2e = emission_ton * gwp
        return round(co2e, 6)
    except Exception as e:
        print(f"    ⚠ Soğutucu hesap hatası [{label}]: {e}")
        return 0.0

def calc_electricity(consumption, unit, ef, label=""):
    """kWh/MWh/GWh → MWh'e çevir → × EF (ton CO2e/MWh)"""
    try:
        val = float(consumption or 0)
        ef  = float(ef or 0)
        if val == 0: return 0.0
        mwh = to_mwh(val, unit, label)
        if mwh is None:
            _warn(f"{label}: bilinmeyen enerji birimi '{unit}', kWh kabul edildi")
            mwh = val / 1000
        return round(mwh * ef, 6)
    except Exception as e:
        print(f"    ⚠ Elektrik hesap hatası [{label}]: {e}")
        return 0.0

def calc_transport(distance, unit, ef, label=""):
    """km/mil/m → km'e çevir → × EF (kg CO2e/tkm) / 1000"""
    try:
        val = float(distance or 0)
        ef  = float(ef or 0)
        if val == 0: return 0.0
        km = to_km(val, unit, label)
        if km is None:
            _warn(f"{label}: bilinmeyen mesafe birimi '{unit}', km kabul edildi")
            km = val
        return round(km * ef / 1000, 6)
    except Exception as e:
        print(f"    ⚠ Taşıma hesap hatası [{label}]: {e}")
        return 0.0

def calc_spend(amount, unit, ef, kur=None, label=""):
    """USD/EUR/GBP/TL → USD'ye çevir → × EF / 1000"""
    try:
        val = float(amount or 0)
        ef  = float(ef or 0)
        if val == 0: return 0.0
        usd = to_usd(val, unit, kur=kur, label=label)
        return round(usd * ef / 1000, 6)
    except Exception as e:
        print(f"    ⚠ Harcama hesap hatası [{label}]: {e}")
        return 0.0

def calc_capital(amount, unit, ef, depreciation, kur=None, label=""):
    """USD/EUR/GBP/TL → USD → × EF / (amortisman × 1000)"""
    try:
        val = float(amount or 0)
        ef  = float(ef or 0)
        dep = float(depreciation or 1) or 1
        if val == 0: return 0.0
        usd = to_usd(val, unit, kur=kur, label=label)
        return round(usd * ef / (dep * 1000), 6)
    except Exception as e:
        print(f"    ⚠ Sermaye hesap hatası [{label}]: {e}")
        return 0.0

def calc_td_loss(consumption, unit, loc_ef, mkt_ef, label=""):
    """T&D kayıp = Tüketim(MWh) × (Pazar EF - Konum EF)"""
    try:
        val    = float(consumption or 0)
        loc_ef = float(loc_ef or 0)
        mkt_ef = float(mkt_ef or 0)
        if val == 0: return 0.0
        mwh = to_mwh(val, unit, label)
        if mwh is None:
            mwh = val / 1000
        return round(mwh * (mkt_ef - loc_ef), 6)
    except Exception as e:
        print(f"    ⚠ T&D hesap hatası [{label}]: {e}")
        return 0.0


# ══════════════════════════════════════════════════════════════════
# EXCEL OKUMA & YAZMA
# ══════════════════════════════════════════════════════════════════

def is_valid_year(val):
    try:
        y = int(float(str(val)))
        return 2000 <= y <= 2100
    except: return False

def read_sheet(wb, sheet_name, header_row=3):
    ws = wb[sheet_name]
    headers = []
    for c in ws[header_row]:
        v = str(c.value or '').strip()
        # Alt başlık satırı varsa (row 4) birleştir
        headers.append(v)

    # Çift başlık varsa (row 3 + row 4) — 1.4 gibi
    # header_row=3 için row 4'ü de kontrol et
    sub_headers = [str(c.value or '').strip() for c in ws[header_row + 1]]
    merged_headers = []
    for h, sh in zip(headers, sub_headers):
        if sh and sh not in h and not is_valid_year(sh):
            merged_headers.append(f"{h} {sh}".strip())
        else:
            merged_headers.append(h)

    data = []
    for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
        vals = list(row[:len(merged_headers)])
        if not any(v is not None for v in vals): continue
        if not is_valid_year(vals[0]): continue
        data.append(vals)

    if not data:
        return pd.DataFrame(columns=merged_headers)

    df = pd.DataFrame(data, columns=merged_headers)
    df['_year'] = df.iloc[:, 0].apply(lambda v: int(float(str(v))))
    return df

def read_sheet_single(wb, sheet_name, header_row=3):
    """Tek başlık satırlı sayfalar için."""
    ws = wb[sheet_name]
    headers = [str(c.value or '').split('\n')[0].strip()
               for c in ws[header_row]]
    data = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = list(row[:len(headers)])
        if not any(v is not None for v in vals): continue
        if not is_valid_year(vals[0]): continue
        data.append(vals)
    if not data:
        return pd.DataFrame(columns=headers)
    df = pd.DataFrame(data, columns=headers)
    df['_year'] = df.iloc[:,0].apply(lambda v: int(float(str(v))))
    return df

thin_s = Side(style="thin", color="BFBFBF")
B_s    = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

def write_grey(cell, value):
    cell.value     = round(value, 4) if value else 0
    cell.font      = Font(name="Arial", size=10, color="595959")
    cell.fill      = PatternFill("solid", start_color="F2F2F2")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = B_s

def imp_c(cell):
    """Turuncu — %95 eşiği içindeki kritik kaynaklar."""
    cell.font      = Font(name="Arial", size=10, color="843C0C")
    cell.fill      = PatternFill("solid", start_color="FCE4D6")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = B_s

def lck(cell):
    """Gri — %95 eşiği dışındaki kaynaklar."""
    cell.font      = Font(name="Arial", size=10, color="595959")
    cell.fill      = PatternFill("solid", start_color="F2F2F2")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = B_s

def write_co2e_column(ws, col_index, results, start_row=4):
    for i, val in enumerate(results):
        write_grey(ws.cell(row=start_row + i, column=col_index), val or 0)


# ══════════════════════════════════════════════════════════════════
# DB HELPERS
# ══════════════════════════════════════════════════════════════════

def get_or_create_period(session, year):
    period = session.query(ReportingPeriod).join(Company).filter(
        Company.name == "UNSPED",
        ReportingPeriod.year == year,
        ReportingPeriod.period == "annual"
    ).first()
    if not period:
        company = session.query(Company).filter_by(name="UNSPED").first()
        if not company:
            raise ValueError("❌ UNSPED şirketi DB'de yok. Önce 'python main.py' çalıştırın.")
        period = ReportingPeriod(company_id=company.id, year=year, period="annual")
        session.add(period)
        session.flush()
        print(f"    ℹ {year} dönemi otomatik oluşturuldu.")
    return period

def clear_year_data(session, period_id, year):
    """
    Belirtilen yılın tüm emisyon verisini siler.
    synchronize_session=False ile SQLAlchemy cache'ini bypass eder.
    """
    deleted = 0
    for Model in [StationaryCombustion, MobileCombustion, Refrigerant,
                  ElectricityConsumption, FreightEmission, EmployeeCommuting,
                  PurchasedGoods, CapitalGoods]:
        n = session.query(Model).filter_by(
            period_id=period_id
        ).delete(synchronize_session=False)
        deleted += n
    session.flush()  # silme işlemini DB'ye gönder
    print(f"    🗑  {year}: {deleted} eski kayıt silindi (period_id={period_id})")


# ══════════════════════════════════════════════════════════════════
# PROCESS FONKSİYONLARI
# ══════════════════════════════════════════════════════════════════

def process_11(wb, session, year_to_period):
    print("  [1.1] Sabit Yakma...")
    df = read_sheet_single(wb, '1.1_Stationary')
    results, totals = [], defaultdict(float)
    for _, r in df.iterrows():
        year = r['_year']
        pid  = year_to_period[year]
        birim  = str(r.get('Birim') or 'L').strip()
        label  = f"{r.get('Emisyon Kaynağı')} ({r.get('Yakıt Türü')})"
        co2e = calc_combustion(
            r.get('Miktar'), birim,
            r.get('NCV'), r.get('EF CO2'),
            r.get('EF CH4'), r.get('EF N2O'),
            r.get('GWP CH4'), r.get('GWP N2O'),
            r.get('Yoğunluk'), label=label
        )
        results.append(co2e); totals[year] += co2e
        session.add(StationaryCombustion(
            period_id=pid,
            emission_source=str(r.get('Emisyon Kaynağı') or ''),
            fuel_type=str(r.get('Yakıt Türü') or ''),
            activity_value=float(r.get('Miktar') or 0),
            activity_unit=birim, co2e_total=co2e
        ))
    write_co2e_column(wb['1.1_Stationary'], 13, results)
    for y, t in totals.items(): print(f"    → {y}: {t:.4f} ton CO2e ✅")
    return totals

def process_12(wb, session, year_to_period):
    print("  [1.2] Hareketli Yakma...")
    df = read_sheet_single(wb, '1.2_Mobile')
    results, totals = [], defaultdict(float)
    for _, r in df.iterrows():
        year = r['_year']
        pid  = year_to_period[year]
        birim = str(r.get('Birim') or 'L').strip()
        label = f"{r.get('Emisyon Kaynağı')} ({r.get('Yakıt Türü')})"
        co2e = calc_combustion(
            r.get('Miktar'), birim,
            r.get('NCV'), r.get('EF CO2'),
            r.get('EF CH4'), r.get('EF N2O'),
            r.get('GWP CH4'), r.get('GWP N2O'),
            r.get('Yoğunluk'), label=label
        )
        results.append(co2e); totals[year] += co2e
        session.add(MobileCombustion(
            period_id=pid,
            emission_source=str(r.get('Emisyon Kaynağı') or ''),
            fuel_type=str(r.get('Yakıt Türü') or ''),
            activity_value=float(r.get('Miktar') or 0),
            activity_unit=birim, co2e_total=co2e
        ))
    write_co2e_column(wb['1.2_Mobile'], 13, results)
    for y, t in totals.items(): print(f"    → {y}: {t:.4f} ton CO2e ✅")
    return totals

def process_14(wb, session, year_to_period):
    """FV × EF × GWP = tCO2e — kg veya ton dinamik."""
    print("  [1.4] Soğutucu Gazlar...")
    ws = wb['1.4_Refrigerants']
    results, totals = [], defaultdict(float)
    # 1.4'te çift başlık var (row 3 grup, row 4 alt) → veri row 5'ten başlar
    for r in range(5, ws.max_row + 1):
        if not is_valid_year(ws.cell(r, 1).value): continue
        year    = int(float(str(ws.cell(r, 1).value)))
        src     = str(ws.cell(r, 2).value or '')
        gas     = str(ws.cell(r, 3).value or '')
        fv      = float(ws.cell(r, 4).value or 0)
        fv_unit = str(ws.cell(r, 5).value or 'kg').strip()
        ef_pct  = float(ws.cell(r, 6).value or 0)
        gwp     = float(ws.cell(r, 10).value or 0)
        pid     = year_to_period.get(year)
        if not pid: continue

        label = f"{src} ({gas})"
        co2e = calc_refrigerant(fv, fv_unit, ef_pct, gwp, label=label)

        # Emisyon değerini col 8'e yaz
        emission = fv * ef_pct
        write_grey(ws.cell(r, 8), emission)
        # Emisyon birim col 9
        uk = unit_key(fv_unit)
        ws.cell(r, 9).value = f"{'kg' if uk=='kg' else 'ton'} {gas}"
        # tCO2e col 11
        write_grey(ws.cell(r, 11), co2e)

        results.append(co2e); totals[year] += co2e
        session.add(Refrigerant(
            period_id=pid, gas_type=gas,
            activity_value=fv, activity_unit=fv_unit,
            gwp=gwp, co2e_total=co2e
        ))
    for y, t in totals.items(): print(f"    → {y}: {t:.4f} ton CO2e ✅")
    return totals

def process_21(wb, session, year_to_period):
    print("  [2.1] Elektrik...")
    df = read_sheet_single(wb, '2.1_Electricity')
    results, totals = [], defaultdict(float)
    for _, r in df.iterrows():
        year  = r['_year']
        pid   = year_to_period[year]
        birim = str(r.get('Birim') or 'kWh').strip()
        ef    = float(r.get('EF') or 0)
        label = f"Elektrik {year}"
        co2e  = calc_electricity(r.get('Tüketim'), birim, ef, label=label)
        mwh   = to_mwh(float(r.get('Tüketim') or 0), birim, label) or 0
        results.append(co2e); totals[year] += co2e
        session.add(ElectricityConsumption(
            period_id=pid, consumption_mwh=mwh,
            emission_factor=ef,
            ef_source=str(r.get('EF Kaynak') or ''),
            co2e_total=co2e
        ))
    write_co2e_column(wb['2.1_Electricity'], 6, results)
    for y, t in totals.items(): print(f"    → {y}: {t:.4f} ton CO2e ✅")
    return totals

def process_freight(wb, session, year_to_period, sheet, label_prefix, cat_default):
    print(f"  [{cat_default}] {label_prefix}...")
    df = read_sheet_single(wb, sheet)
    results, totals = [], defaultdict(float)
    for _, r in df.iterrows():
        year  = r['_year']
        pid   = year_to_period[year]
        birim = str(r.get('Birim') or 'km').strip()
        label = str(r.get('Emisyon Kaynağı') or '')
        co2e  = calc_transport(r.get('Mesafe'), birim,
                               r.get('EF'), label=label)
        results.append(co2e); totals[year] += co2e
        session.add(FreightEmission(
            period_id=pid,
            category_code=str(r.get('Kategori') or cat_default),
            emission_source=label,
            transport_type=str(r.get('Taşıma Tipi') or ''),
            vehicle_type=str(r.get('Araç Tipi') or ''),
            activity_value=float(r.get('Mesafe') or 0),
            activity_unit='km',
            emission_factor=float(r.get('EF') or 0),
            co2e_total=co2e
        ))
    write_co2e_column(wb[sheet], 9, results)
    for y, t in totals.items(): print(f"    → {y}: {t:.4f} ton CO2e ✅")
    return totals

def process_33(wb, session, year_to_period):
    print("  [3.3] Personel Ulaşım...")
    df = read_sheet_single(wb, '3.3_Commuting')
    results, totals = [], defaultdict(float)
    for _, r in df.iterrows():
        year  = r['_year']
        pid   = year_to_period[year]
        birim = str(r.get('Birim') or 'L').strip()
        label = f"{r.get('Emisyon Kaynağı')} ({r.get('Yakıt Türü')})"
        co2e = calc_combustion(
            r.get('Miktar'), birim,
            r.get('NCV'), r.get('EF CO2'),
            r.get('EF CH4'), r.get('EF N2O'),
            r.get('GWP CH4'), r.get('GWP N2O'),
            r.get('Yoğunluk'), label=label
        )
        results.append(co2e); totals[year] += co2e
        session.add(EmployeeCommuting(
            period_id=pid,
            category_code=str(r.get('Kategori') or ''),
            emission_source=str(r.get('Emisyon Kaynağı') or ''),
            fuel_type=str(r.get('Yakıt Türü') or ''),
            activity_value=float(r.get('Miktar') or 0),
            activity_unit=birim, co2e_total=co2e
        ))
    write_co2e_column(wb['3.3_Commuting'], 14, results)
    for y, t in totals.items(): print(f"    → {y}: {t:.4f} ton CO2e ✅")
    return totals

def process_35(wb, session, year_to_period):
    print("  [3.5] Uçak & Konaklama...")
    df = read_sheet_single(wb, '3.5_FlightsHotels')
    results, totals = [], defaultdict(float)
    for _, r in df.iterrows():
        year  = r['_year']
        pid   = year_to_period[year]
        tl    = float(r.get('Miktar (TL)') or 0)
        kur   = float(r.get('Dolar Kuru') or 1) or 1
        ef    = float(r.get('EF') or 0)
        label = str(r.get('Kalem') or '')
        # TL → USD
        usd   = to_usd(tl, 'TL', kur=kur, label=label)
        # USD ara değeri col 6'ya yaz
        write_grey(wb['3.5_FlightsHotels'].cell(
            row=4 + list(df.index).index(r.name), column=6), usd)
        co2e = round(usd * ef / 1000, 6)
        results.append(co2e); totals[year] += co2e
        session.add(PurchasedGoods(
            period_id=pid,
            category_code=str(r.get('Kategori') or ''),
            item_name=label,
            activity_value=tl, activity_unit='TL',
            emission_factor=ef,
            ef_source='SupplyChainGHG / DEFRA',
            co2e_total=co2e
        ))
    write_co2e_column(wb['3.5_FlightsHotels'], 8, results)
    for y, t in totals.items(): print(f"    → {y}: {t:.4f} ton CO2e ✅")
    return totals

def process_purchased(wb, session, year_to_period, sheet, co2e_col, label_prefix):
    print(f"  {label_prefix}...")
    df = read_sheet_single(wb, sheet)
    results, totals = [], defaultdict(float)
    for _, r in df.iterrows():
        year  = r['_year']
        pid   = year_to_period[year]
        birim = str(r.get('Birim') or 'USD').strip()
        # Miktar sütununu bul — sayısal olan ilk veri sütunu
        skip = {'_year','Yıl','Kategori','Birim','EF','EF Kaynak','Not',
                'Ürün Adı','Atık Türü','Ekipman Adı','Hizmet Adı','Varlık Adı'}
        amt_key = [c for c in df.columns if c not in skip
                   and c not in df.columns[:3]
                   and r.get(c) is not None
                   and str(r.get(c)).replace('.','').replace('-','').isdigit()]
        amt = float(r.get(amt_key[0]) if amt_key else 0)
        label = str(r.get(list(df.columns)[2]) or '')
        co2e = calc_spend(amt, birim, r.get('EF'), label=label)
        results.append(co2e); totals[year] += co2e
        session.add(PurchasedGoods(
            period_id=pid,
            category_code=str(r.get('Kategori') or ''),
            item_name=label,
            activity_value=amt, activity_unit=birim,
            emission_factor=float(r.get('EF') or 0),
            ef_source=str(r.get('EF Kaynak') or ''),
            co2e_total=co2e
        ))
    write_co2e_column(wb[sheet], co2e_col, results)
    for y, t in totals.items(): print(f"    → {y}: {t:.4f} ton CO2e ✅")
    return totals

def process_42(wb, session, year_to_period):
    print("  [4.2] Sermaye Varlıkları...")
    df = read_sheet_single(wb, '4.2_Capital')
    results, totals = [], defaultdict(float)
    for _, r in df.iterrows():
        year  = r['_year']
        pid   = year_to_period[year]
        birim = str(r.get('Birim') or '2021 USD').strip()
        label = str(r.get('Varlık Adı') or '')
        co2e  = calc_capital(
            r.get('Tutar (2021 USD)'), birim,
            r.get('EF'), r.get('Amortisman (Yıl)'),
            label=label
        )
        results.append(co2e); totals[year] += co2e
        session.add(CapitalGoods(
            period_id=pid,
            category_code=str(r.get('Kategori') or ''),
            asset_name=label,
            activity_value=float(r.get('Tutar (2021 USD)') or 0),
            activity_unit='2021 USD',
            depreciation_years=int(float(r.get('Amortisman (Yıl)') or 1)),
            emission_factor=float(r.get('EF') or 0),
            ef_source=str(r.get('EF Kaynak') or ''),
            co2e_total=co2e
        ))
    write_co2e_column(wb['4.2_Capital'], 9, results)
    for y, t in totals.items(): print(f"    → {y}: {t:.4f} ton CO2e ✅")
    return totals

def process_61(wb, session, year_to_period):
    print("  [6.1] T&D Kayıpları...")
    df = read_sheet_single(wb, '6.1_TDLosses')
    results, totals = [], defaultdict(float)
    for _, r in df.iterrows():
        year   = r['_year']
        pid    = year_to_period[year]
        birim  = str(r.get('Birim') or 'kWh').strip()
        loc_ef = float(r.get('Konum EF') or 0)
        mkt_ef = float(r.get('Pazar EF') or 0)
        label  = f"T&D {year}"
        co2e   = calc_td_loss(r.get('Tüketim'), birim,
                              loc_ef, mkt_ef, label=label)
        mwh    = to_mwh(float(r.get('Tüketim') or 0), birim, label) or 0
        # T&D fark EF col 6
        write_grey(wb['6.1_TDLosses'].cell(
            row=4 + list(df.index).index(r.name), column=6),
            round(mkt_ef - loc_ef, 6))
        results.append(co2e); totals[year] += co2e
        session.add(ElectricityConsumption(
            period_id=pid,
            consumption_mwh=mwh,
            emission_factor=round(mkt_ef - loc_ef, 6),
            ef_source='T&D Kayıp Faktörü',
            co2e_total=co2e
        ))
    write_co2e_column(wb['6.1_TDLosses'], 7, results)
    for y, t in totals.items(): print(f"    → {y}: {t:.4f} ton CO2e ✅")
    return totals


# ══════════════════════════════════════════════════════════════════
# ÖZET SAYFALAR
# ══════════════════════════════════════════════════════════════════

def write_impact(wb, scope_totals):
    """
    IMPACT_ANALYSIS sayfasını yazar.
    SADECE Kapsam 3 kategorileri (3.1 → 6.1) gösterilir.
    Tüm kaynaklar listelenir — %95 eşiği işaretlenir ama hiçbiri atlanmaz.
    """
    latest_year = max(scope_totals.keys())
    totals      = scope_totals[latest_year]
    ws          = wb['IMPACT_ANALYSIS']

    # Sadece Kapsam 3 kategorileri
    s3_sources = [
        ('3.1', 'Hammadde Sevkiyatı'),
        ('3.2', 'Ürün Sevkiyatı'),
        ('3.3', 'Personel Ulaşım'),
        ('3.4', 'İş Seyahati - Karayolu'),
        ('3.5', 'Uçak & Konaklama'),
        ('4.1', 'Satın Alınan Malzemeler'),
        ('4.2', 'Sermaye Varlıkları'),
        ('4.3', 'Atık Bertarafı'),
        ('4.4', 'Kiralanan Ekipmanlar'),
        ('4.5', 'Hizmet Alımları'),
        ('6.1', 'T&D Kayıpları'),
    ]

    # Kapsam 3 toplamı
    s3_total = totals.get('Kapsam 3 Top.', 0) or 0
    if s3_total == 0:
        s3_total = sum(totals.get(cat, 0) or 0 for cat, _ in s3_sources)

    # Sıralamak için değerleri al
    sources_with_val = [
        ('Kapsam 3', cat, label, totals.get(cat, 0) or 0)
        for cat, label in s3_sources
    ]
    sources_sorted = sorted(sources_with_val, key=lambda x: x[3], reverse=True)

    # Mevcut merge'leri kaldır, sonra temizle
    for merge in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge))
    for row in ws.iter_rows(min_row=4, max_row=40):
        for cell in row:
            try:
                cell.value = None
                cell.fill  = PatternFill("solid", start_color="FFFFFF")
            except: pass

    cumulative = 0
    active_row = 4
    threshold_written = False

    for scope, cat, label, val in sources_sorted:
        if s3_total > 0:
            pct = round(val / s3_total * 100, 2)
        else:
            pct = 0
        cumulative = round(cumulative + pct, 2)
        above_95   = cumulative <= 95.0

        # %95 eşiği geçildiyse ayraç yaz
        if not above_95 and not threshold_written:
            ws.merge_cells(f"A{active_row}:G{active_row}")
            m = ws.cell(row=active_row, column=1)
            m.value = (f"▲ %95 EŞİĞİ ({latest_year}) — "
                       f"Aşağıdaki kaynaklar Kapsam 3'ün %5'inden az")
            m.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            m.fill      = PatternFill("solid", start_color="843C0C")
            m.alignment = Alignment(horizontal="center", vertical="center")
            active_row += 1
            threshold_written = True

        for ci, v in enumerate([
            active_row - 3, scope, cat, label,
            round(val, 4), f"{pct}%", f"{cumulative}%"
        ], 1):
            cell = ws.cell(row=active_row, column=ci, value=v)
            if above_95:
                imp_c(cell)
            else:
                lck(cell)
        active_row += 1

    # Eğer threshold hiç yazılmadıysa (hepsi %95 içindeyse) sona yaz
    if not threshold_written:
        ws.merge_cells(f"A{active_row}:G{active_row}")
        m = ws.cell(row=active_row, column=1)
        m.value = f"▲ TÜM KAYNAKLAR %95 EŞİĞİ İÇİNDE ({latest_year})"
        m.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        m.fill      = PatternFill("solid", start_color="1F4E79")
        m.alignment = Alignment(horizontal="center", vertical="center")

    print(f"    ✅ IMPACT_ANALYSIS güncellendi ({latest_year} | Kapsam 3 toplam: {s3_total:.2f} ton CO2e)")


def merge_totals(all_results):
    all_years = set()
    for d in all_results.values(): all_years.update(d.keys())
    scope_totals = {}
    for year in sorted(all_years):
        g = {k: all_results.get(k, {}).get(year, 0) for k in
             ['1.1','1.2','1.4','2.1','3.1','3.2','3.3',
              '3.4','3.5','4.1','4.2','4.3','4.4','4.5','6.1']}
        s1 = g['1.1']+g['1.2']+g['1.4']
        s2 = g['2.1']
        s3 = sum(g[k] for k in ['3.1','3.2','3.3','3.4','3.5',
                                  '4.1','4.2','4.3','4.4','4.5','6.1'])
        scope_totals[year] = {**g,
            'Kapsam 1 Top.': s1, 'Kapsam 2 Top.': s2,
            'Kapsam 3 Top.': s3, 'GENEL TOPLAM': s1+s2+s3}
    return scope_totals

def write_per_capita(wb, scope_totals):
    """
    PER_CAPITA sayfasını otomatik doldurur.
    PERSONNEL sayfasından yıl + lokasyon + personel sayısını çeker.
    Kullanıcı sadece PERSONNEL'e: Yıl, Lokasyon Adı, Personel Sayısı girer.
    CO2e değerleri scope_totals'tan otomatik hesaplanır.
    """
    ws_pc  = wb['PER_CAPITA']
    ws_per = wb['PERSONNEL']

    # PERSONNEL sayfasını oku — {year: [{location, headcount}]}
    personnel_by_year = {}
    for row in ws_per.iter_rows(min_row=4, values_only=True):
        if not row[0] or not row[1] or not row[2]: continue
        try:
            year = int(float(str(row[0])))
            if not (2000 <= year <= 2100): continue
            loc  = str(row[1]).strip()
            cnt  = int(float(str(row[2])))
            if year not in personnel_by_year:
                personnel_by_year[year] = []
            personnel_by_year[year].append({'location': loc, 'headcount': cnt})
        except: continue

    if not personnel_by_year:
        print("    ⚠ PERSONNEL sayfasında veri bulunamadı.")
        return

    # Merge'leri kaldır, sonra temizle
    for merge in list(ws_pc.merged_cells.ranges):
        ws_pc.unmerge_cells(str(merge))
    for row in ws_pc.iter_rows(min_row=4, max_row=100):
        for cell in row:
            try:
                cell.value = None
                cell.fill  = PatternFill("solid", start_color="FFFFFF")
            except: pass

    current_row = 4
    for year in sorted(personnel_by_year.keys()):
        if year not in scope_totals:
            print(f"    ⚠ {year} için emisyon verisi yok, PER_CAPITA atlandı.")
            continue

        personnel = personnel_by_year[year]
        totals    = scope_totals[year]
        s1        = totals.get('Kapsam 1 Top.', 0) or 0
        s2        = totals.get('Kapsam 2 Top.', 0) or 0
        s3        = totals.get('Kapsam 3 Top.', 0) or 0
        grand     = s1 + s2 + s3
        total_head = sum(p['headcount'] for p in personnel)

        if total_head == 0: continue

        for p in personnel:
            ratio     = p['headcount'] / total_head
            loc_s1    = round(s1 * ratio, 4)
            loc_s2    = round(s2 * ratio, 4)
            loc_s3    = round(s3 * ratio, 4)
            loc_total = round(loc_s1 + loc_s2 + loc_s3, 4)
            per_cap   = round(loc_total / p['headcount'], 6) if p['headcount'] else 0

            ws_pc.row_dimensions[current_row].height = 20
            for ci, val in enumerate([
                year, p['location'], p['headcount'],
                loc_total, loc_s1, loc_s2, loc_s3, per_cap
            ], 1):
                cell = ws_pc.cell(row=current_row, column=ci, value=val)
                cell.font      = Font(name="Arial", size=10, color="00008B")
                cell.fill      = PatternFill("solid", start_color="D9E1F2")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = B_s
            current_row += 1

        # Yıl toplamı satırı
        grand_cap = round(grand / total_head, 6) if total_head else 0
        ws_pc.row_dimensions[current_row].height = 20
        for ci, val in enumerate([
            year, f'ŞİRKET TOPLAMI ({year})', total_head,
            round(grand, 4), round(s1, 4),
            round(s2, 4), round(s3, 4), grand_cap
        ], 1):
            cell = ws_pc.cell(row=current_row, column=ci, value=val)
            cell.font      = Font(name="Arial", size=10, bold=True, color="7B3F00")
            cell.fill      = PatternFill("solid", start_color="FFF2CC")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = B_s
        current_row += 2  # boş satır

        print(f"    ✅ {year}: {len(personnel)} lokasyon | "
              f"K1={s1:.2f} K2={s2:.2f} K3={s3:.2f} | "
              f"Toplam={grand:.2f} ton | {grand_cap:.4f} ton/kişi")


def write_totals(wb, scope_totals):
    """
    TOTAL_EMISSIONS sayfasını yazar.
    Tüm kategoriler yazılır — hiçbir şey atlanmaz.
    Değişim % sütunu son iki yıl arasında hesaplanır.
    """
    ws = wb['TOTAL_EMISSIONS']
    cat_rows = {
        '1.1':5,'1.2':6,'1.4':7,'Kapsam 1 Top.':8,
        '2.1':11,'Kapsam 2 Top.':12,
        '3.1':15,'3.2':16,'3.3':17,'3.4':18,'3.5':19,
        '4.1':20,'4.2':21,'4.3':22,'4.4':23,'4.5':24,
        '6.1':25,'Kapsam 3 Top.':26,'GENEL TOPLAM':28,
    }
    sorted_years = sorted(scope_totals.keys())

    for i, year in enumerate(sorted_years):
        totals = scope_totals[year]
        col = 3 + i
        ws.cell(row=3, column=col).value = f"{year} (ton CO2e)"
        for cat, val in totals.items():
            row = cat_rows.get(cat)
            if not row: continue
            write_grey(ws.cell(row, col), round(val, 4) if val else 0)

    # Değişim % — son iki yıl arasında
    if len(sorted_years) >= 2:
        prev_year = sorted_years[-2]
        last_year = sorted_years[-1]
        prev_totals = scope_totals[prev_year]
        last_totals = scope_totals[last_year]
        chg_col = 3 + len(sorted_years)  # son yıl sütunundan sonra
        ws.cell(row=3, column=chg_col).value = f"Değişim % ({prev_year}→{last_year})"
        for cat, row in cat_rows.items():
            prev_val = prev_totals.get(cat, 0) or 0
            last_val = last_totals.get(cat, 0) or 0
            if prev_val != 0:
                chg = round((last_val - prev_val) / prev_val * 100, 1)
                ws.cell(row=row, column=chg_col).value = f"{'+' if chg>0 else ''}{chg}%"
            else:
                ws.cell(row=row, column=chg_col).value = "—"

    print("    ✅ TOTAL_EMISSIONS güncellendi.")


# ══════════════════════════════════════════════════════════════════
# ANA FONKSİYON
# ══════════════════════════════════════════════════════════════════

def run_import(excel_path):
    conversion_warnings.clear()
    print(f"\n{'='*58}")
    print(f"  UNSPED KARBON IMPORTER v4")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Dosya: {excel_path}")
    print(f"{'='*58}\n")

    wb = load_workbook(excel_path)

    with SessionLocal() as session:
        # Tüm sayfaları tara — hangi yıllar var
        all_years = set()
        for sname in ['1.1_Stationary','1.2_Mobile','1.4_Refrigerants',
                       '2.1_Electricity','3.1_Freight','3.2_ProductShipment',
                       '3.3_Commuting','3.4_BusinessTravel','3.5_FlightsHotels',
                       '4.1_Purchased','4.2_Capital','4.3_Waste',
                       '4.4_LeasedEquipment','4.5_Services','6.1_TDLosses']:
            try:
                ws = wb[sname]
                for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
                    if row and is_valid_year(row[0]):
                        all_years.add(int(float(str(row[0]))))
            except: pass

        if not all_years:
            raise ValueError("❌ Excel'de geçerli yıl verisi bulunamadı.")

        print(f"  Bulunan yıllar: {sorted(all_years)}\n")

        # Period al + temizle
        year_to_period = {}
        for year in sorted(all_years):
            period = get_or_create_period(session, year)
            year_to_period[year] = period.id
        # Önce hepsini commit et (period oluşturma)
        session.flush()
        # Sonra her yılın verisini temizle
        for year, pid in year_to_period.items():
            clear_year_data(session, pid, year)
        session.flush()  # silmeleri de gönder

        # Sayfaları işle
        print("KAPSAM 1")
        r11 = process_11(wb, session, year_to_period)
        r12 = process_12(wb, session, year_to_period)
        r14 = process_14(wb, session, year_to_period)

        print("\nKAPSAM 2")
        r21 = process_21(wb, session, year_to_period)

        print("\nKAPSAM 3")
        r31 = process_freight(wb, session, year_to_period, '3.1_Freight',         'Hammadde Sevkiyatı','3.1')
        r32 = process_freight(wb, session, year_to_period, '3.2_ProductShipment', 'Ürün Sevkiyatı',   '3.2')
        r33 = process_33(wb, session, year_to_period)
        r34 = process_freight(wb, session, year_to_period, '3.4_BusinessTravel',  'İş Seyahati',      '3.4')
        r35 = process_35(wb, session, year_to_period)
        r41 = process_purchased(wb, session, year_to_period, '4.1_Purchased',      8, '[4.1] Satın Alınan')
        r42 = process_42(wb, session, year_to_period)
        r43 = process_purchased(wb, session, year_to_period, '4.3_Waste',          8, '[4.3] Atık')
        r44 = process_purchased(wb, session, year_to_period, '4.4_LeasedEquipment',8, '[4.4] Kiralanan')
        r45 = process_purchased(wb, session, year_to_period, '4.5_Services',       8, '[4.5] Hizmetler')
        r61 = process_61(wb, session, year_to_period)

        all_results = {
            '1.1':r11,'1.2':r12,'1.4':r14,'2.1':r21,
            '3.1':r31,'3.2':r32,'3.3':r33,'3.4':r34,'3.5':r35,
            '4.1':r41,'4.2':r42,'4.3':r43,'4.4':r44,'4.5':r45,'6.1':r61,
        }
        scope_totals = merge_totals(all_results)

        # Audit log
        for year, totals in scope_totals.items():
            session.add(AuditLog(
                action="excel_import", scope="all",
                period_id=year_to_period[year], status="success",
                notes=(f"{year} | "
                       f"K1={totals['Kapsam 1 Top.']:.2f} "
                       f"K2={totals['Kapsam 2 Top.']:.2f} "
                       f"K3={totals['Kapsam 3 Top.']:.2f} "
                       f"Toplam={totals['GENEL TOPLAM']:.2f} ton CO2e")
            ))
        session.commit()
        print(f"\n  DB'ye kaydedildi.")

    # Özet sayfalar
    print("\nÖZET SAYFALAR")
    write_totals(wb, scope_totals)
    write_impact(wb, scope_totals)
    write_per_capita(wb, scope_totals)
    wb.save(excel_path)

    # Dönüşüm özeti
    if conversion_warnings:
        print(f"\n{'─'*55}")
        print(f"  BİRİM DÖNÜŞÜM ÖZETİ ({len(conversion_warnings)} dönüşüm)")
        print(f"{'─'*55}")
        for w in conversion_warnings:
            print(f"  • {w}")

    print(f"\n{'='*58}")
    print(f"  IMPORT TAMAMLANDI")
    for year in sorted(scope_totals.keys()):
        t = scope_totals[year]
        print(f"\n  {year}")
        print(f"    Kapsam 1 : {t['Kapsam 1 Top.']:.4f} ton CO2e")
        print(f"    Kapsam 2 : {t['Kapsam 2 Top.']:.4f} ton CO2e")
        print(f"    Kapsam 3 : {t['Kapsam 3 Top.']:.4f} ton CO2e")
        print(f"    TOPLAM   : {t['GENEL TOPLAM']:.4f} ton CO2e")
    print(f"{'='*58}\n")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else \
           "data/UNSPED_Karbon_Veri_Girisi_v3.xlsx"
    run_import(path)