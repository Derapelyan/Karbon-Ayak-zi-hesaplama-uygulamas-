"""
UNSPED Karbon Ayak İzi Dashboard v3
python dashboard.py
"""
import sys, os, threading, io, json, re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

# ── Renkler & Fontlar ──────────────────────────────────────────────
C = {
    "bg":"#F5F5F0","sidebar":"#1F4E79","sidebar_hl":"#2E75B6",
    "white":"#FFFFFF","text":"#2C2C2A","muted":"#888780","border":"#D3D1C7",
    "blue":"#1F4E79","green":"#1D9E75","amber":"#BA7517","red":"#A32D2D",
    "s1":"#2E75B6","s2":"#1D9E75","s3":"#BA7517","card":"#FFFFFF",
    "success":"#E2EFDA","warning":"#FFF2CC","danger":"#FCEBEB","impact":"#FCE4D6",
}
FONT       = ("Segoe UI", 10)
FONT_BOLD  = ("Segoe UI", 10, "bold")
FONT_TITLE = ("Segoe UI", 14, "bold")
FONT_LARGE = ("Segoe UI", 22, "bold")
FONT_SMALL = ("Segoe UI", 9)
FONT_MONO  = ("Consolas", 9)

DISPLAY_YEARS = 3

# ── EVÇED Referans ────────────────────────────────────────────────
EVCED_FACTORS = {
    2019:{"uretim":0.481,"tuketim":0.488,"kaynak":"ETKB EVÇED 2019"},
    2020:{"uretim":0.452,"tuketim":0.458,"kaynak":"ETKB EVÇED 2020"},
    2021:{"uretim":0.434,"tuketim":0.439,"kaynak":"ETKB EVÇED 2021"},
    2022:{"uretim":0.442,"tuketim":0.442,"kaynak":"ETKB EVÇED 2022"},
    2023:{"uretim":0.434,"tuketim":0.434,"kaynak":"ETKB EVÇED 2023"},
}
EVCED_URL = "https://enerji.gov.tr/evced-cevre-ve-iklim-iklim-degisikligi-emisyon-faktorleri"

GWP_TABLE = [
    ("CO2","Karbondioksit",1,"IPCC AR6 2021"),
    ("CH4","Metan",29.8,"IPCC AR6 2021"),
    ("N2O","Diazotmonoksit",273,"IPCC AR6 2021"),
    ("SF6","Kükürt hekzaflorür",25200,"IPCC AR6 2021"),
    ("R134A","HFC-134a",1526,"IPCC AR6 2021"),
    ("R32","HFC-32",771,"IPCC AR6 2021"),
    ("R410A","HFC-410A",2088,"IPCC AR6 2021"),
    ("R407C","HFC-407C",1774,"IPCC AR6 2021"),
    ("R404A","HFC-404A",3922,"IPCC AR6 2021"),
    ("R507A","HFC-507A",3985,"IPCC AR6 2021"),
    ("R22","HCFC-22",1960,"IPCC AR6 2021"),
    ("R290","Propan",0,"IPCC AR6 2021"),
    ("FM200","HFC-227ea",3220,"IPCC AR6 2021"),
    ("HFC236fa","HFC-236fa",8060,"IPCC AR6 2021"),
    ("R448A","HFC-448A",1387,"IPCC AR6 2021"),
    ("R454B","HFC-454B",466,"IPCC AR6 2021"),
]

EF_TABLE = [
    ("Doğal Gaz","m³",56.1,0.005,0.0001,48.0,0.000717,"IPCC 2006 GL"),
    ("Dizel","L",74.1,0.010,0.0006,43.0,0.000832,"IPCC 2006 GL"),
    ("Benzin","L",69.3,0.010,0.0006,44.3,0.000745,"IPCC 2006 GL"),
    ("LPG","L",63.1,0.010,0.0006,47.3,0.000540,"IPCC 2006 GL"),
    ("Fuel Oil","L",77.4,0.010,0.0006,40.4,0.000890,"IPCC 2006 GL"),
    ("Kömür","ton",101.0,0.010,0.0015,11.9,1.0,"IPCC 2006 GL"),
]

NCV_TABLE = [
    ("Doğal Gaz","m³",48.0,"TJ/Gg",0.000717,"ton/m³","IPCC 2006 GL"),
    ("Dizel","L",43.0,"TJ/Gg",0.000832,"ton/L","IPCC 2006 GL"),
    ("Benzin","L",44.3,"TJ/Gg",0.000745,"ton/L","IPCC 2006 GL"),
    ("LPG","L",47.3,"TJ/Gg",0.000540,"ton/L","IPCC 2006 GL"),
    ("Fuel Oil","L",40.4,"TJ/Gg",0.000890,"ton/L","IPCC 2006 GL"),
    ("Kömür","ton",11.9,"TJ/Gg",1.0,"ton/ton","IPCC 2006 GL"),
]


# ══════════════════════════════════════════════════════════════════
# LOG WRITER
# ══════════════════════════════════════════════════════════════════

class LogWriter:
    def __init__(self, widget):
        self.widget = widget
    def write(self, text):
        self.widget.after(0, lambda t=text: self._append(t))
    def _append(self, text):
        try:
            self.widget.configure(state="normal")
            self.widget.insert("end", text)
            self.widget.see("end")
        except: pass
    def flush(self): pass


# ══════════════════════════════════════════════════════════════════
# GÖRÜNTÜLEME — Son N yıl filtresi
# ══════════════════════════════════════════════════════════════════

def filter_display_years(years_data):
    if len(years_data) <= DISPLAY_YEARS:
        return years_data
    keys = sorted(years_data.keys())[-DISPLAY_YEARS:]
    return {k: years_data[k] for k in keys}


# ══════════════════════════════════════════════════════════════════
# DB YARDIMCILARI — Direkt tablo okuma, regex yok
# ══════════════════════════════════════════════════════════════════

def db_get_scope_totals():
    """DB tablolarından direkt okur. Döner: {yıl: {k1,k2,k3,total,cats}}"""
    try:
        from db.connection import SessionLocal
        from db.models import (ReportingPeriod, StationaryCombustion,
                               MobileCombustion, Refrigerant,
                               ElectricityConsumption, FreightEmission,
                               EmployeeCommuting, PurchasedGoods, CapitalGoods)
        results = {}
        with SessionLocal() as s:
            for p in s.query(ReportingPeriod).order_by(ReportingPeriod.year).all():
                y = str(p.year)
                cats = {k:0.0 for k in ["1.1","1.2","1.4","2.1","3.1","3.2",
                                         "3.3","3.4","3.5","4.1","4.2","4.3",
                                         "4.4","4.5","6.1"]}
                for r in s.query(StationaryCombustion).filter_by(period_id=p.id).all():
                    cats["1.1"] += (r.co2e_total or 0)
                for r in s.query(MobileCombustion).filter_by(period_id=p.id).all():
                    cats["1.2"] += (r.co2e_total or 0)
                for r in s.query(Refrigerant).filter_by(period_id=p.id).all():
                    cats["1.4"] += (r.co2e_total or 0)
                for r in s.query(ElectricityConsumption).filter_by(period_id=p.id).all():
                    cats["2.1"] += (r.co2e_total or 0)
                for r in s.query(FreightEmission).filter_by(period_id=p.id).all():
                    cat = (r.category_code or "3.1")[:3]
                    if cat in cats: cats[cat] += (r.co2e_total or 0)
                for r in s.query(EmployeeCommuting).filter_by(period_id=p.id).all():
                    cat = (r.category_code or "3.3")[:3]
                    if cat in cats: cats[cat] += (r.co2e_total or 0)
                for r in s.query(PurchasedGoods).filter_by(period_id=p.id).all():
                    cat = (r.category_code or "4.1")[:3]
                    if cat in cats: cats[cat] += (r.co2e_total or 0)
                for r in s.query(CapitalGoods).filter_by(period_id=p.id).all():
                    cats["4.2"] += (r.co2e_total or 0)

                k1 = cats["1.1"]+cats["1.2"]+cats["1.4"]
                k2 = cats["2.1"]
                k3 = sum(cats[k] for k in ["3.1","3.2","3.3","3.4","3.5",
                                            "4.1","4.2","4.3","4.4","4.5","6.1"])
                total = k1+k2+k3
                if total > 0 or any(v > 0 for v in cats.values()):
                    results[y] = {
                        "k1":round(k1,4), "k2":round(k2,4),
                        "k3":round(k3,4), "total":round(total,4),
                        "cats":{k:round(v,4) for k,v in cats.items()}
                    }
                    # Debug: beklenmedik yüksek değer varsa logla
                    if total > 5000:
                        print(f"UYARI: {y} için toplam çok yüksek: {total:.1f} ton CO2e")
                        print(f"  K1={k1:.1f} K2={k2:.1f} K3={k3:.1f}")
                        print(f"  Cats: {cats}")
        return results
    except Exception as e:
        print(f"db_get_scope_totals hata: {e}")
        return {}

def db_get_category_breakdown():
    data = db_get_scope_totals()
    return {y: v["cats"] for y, v in data.items()}

def db_get_revisions():
    try:
        from db.connection import SessionLocal
        from db.models import AuditLog
        results = []
        year_counters = {}
        with SessionLocal() as s:
            logs = s.query(AuditLog).filter_by(
                action="excel_import", status="success"
            ).order_by(AuditLog.id.asc()).all()
            for log in logs:
                notes = log.notes or ""
                for m in re.finditer(
                    r'(\d{4})\s*\|\s*K1=([\d.]+)\s*K2=([\d.]+)\s*K3=([\d.]+)\s*Toplam=([\d.]+)',
                    notes
                ):
                    y = m.group(1)
                    year_counters[y] = year_counters.get(y,0)+1
                    results.append({
                        "id":f"{y}-R{year_counters[y]}", "year":y,
                        "ts":str(log.created_at)[:16] if log.created_at else "—",
                        "k1":float(m.group(2)),"k2":float(m.group(3)),
                        "k3":float(m.group(4)),"total":float(m.group(5))
                    })
        return list(reversed(results))
    except: return []

def db_get_personnel():
    """PERSONNEL sayfasından (veya DB'den) personel verisi döner."""
    try:
        from db.connection import SessionLocal
        from db.models import ReportingPeriod
        # Şimdilik period'ları döner, ileride PERSONNEL tablosu eklenebilir
        return {}
    except: return {}


# ══════════════════════════════════════════════════════════════════
# YARDIMCI UI
# ══════════════════════════════════════════════════════════════════

def make_treeview(parent, columns, col_widths, height=12):
    frame = tk.Frame(parent, bg=C["bg"])
    frame.pack(fill="both", expand=True)
    style = ttk.Style()
    style.configure("Custom.Treeview", font=FONT, rowheight=26)
    style.configure("Custom.Treeview.Heading", font=FONT_BOLD)
    tree = ttk.Treeview(frame, columns=columns, show="headings",
                        height=height, style="Custom.Treeview")
    for col, w in zip(columns, col_widths):
        tree.heading(col, text=col)
        tree.column(col, width=w, anchor="center")
    sb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=sb.set)
    tree.pack(side="left", fill="both", expand=True)
    sb.pack(side="right", fill="y")
    return tree

def card(parent, title=None, pady=12):
    outer = tk.Frame(parent, bg=C["card"],
                     highlightbackground=C["border"], highlightthickness=1)
    outer.pack(fill="x", pady=(0,pady))
    inner = tk.Frame(outer, bg=C["card"], padx=16, pady=12)
    inner.pack(fill="both")
    if title:
        tk.Label(inner, text=title, font=FONT_BOLD,
                 bg=C["card"], fg=C["text"]).pack(anchor="w", pady=(0,6))
    return inner

def section_label(parent, text, color=None):
    tk.Label(parent, text=text, font=FONT_BOLD,
             bg=C["bg"], fg=color or C["text"]).pack(anchor="w", pady=(12,4))

def no_data_label(parent, msg="Henüz veri yok.\nÖnce Excel'i import edin."):
    tk.Label(parent, text=msg, font=FONT,
             bg=C["bg"], fg=C["muted"]).pack(expand=True)


# ══════════════════════════════════════════════════════════════════
# ANA UYGULAMA
# ══════════════════════════════════════════════════════════════════

class UNSPEDDashboard(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("UNSPED Karbon Ayak İzi v3")
        self.geometry("1380x870")
        self.minsize(1100,700)
        self.configure(bg=C["bg"])
        self.excel_path = tk.StringVar(value="")
        self.status_var = tk.StringVar(value="Hazır")
        self._build_ui()

    def _build_ui(self):
        self.sidebar = tk.Frame(self, bg=C["sidebar"], width=235)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self.content = tk.Frame(self, bg=C["bg"])
        self.content.pack(side="left", fill="both", expand=True)
        self._build_sidebar()
        self._build_statusbar()
        self._show_page("home")

    def _build_sidebar(self):
        lf = tk.Frame(self.sidebar, bg=C["sidebar"], pady=20)
        lf.pack(fill="x")
        tk.Label(lf, text="🌍", font=("Segoe UI",28), bg=C["sidebar"], fg="white").pack()
        tk.Label(lf, text="UNSPED", font=("Segoe UI",13,"bold"),
                 bg=C["sidebar"], fg="white").pack()
        tk.Label(lf, text="Karbon Dashboard v3", font=("Segoe UI",8),
                 bg=C["sidebar"], fg="#9FC5E8").pack()
        ttk.Separator(self.sidebar).pack(fill="x", padx=16, pady=4)
        self.nav_btns = {}
        menu = [("home","🏠","Ana Sayfa"),("import","📥","Veri İmport"),
                ("update","🔄","Faktör Güncelle"),("results","📊","Sonuçlar & Grafik"),
                ("export","📄","Word Rapor Export"),
                ("refs","📚","Referans Tablolar"),("log","📋","Log & Revizyonlar")]
        for page, icon, label in menu:
            btn = tk.Button(self.sidebar, text=f"  {icon}  {label}",
                font=FONT, bg=C["sidebar"], fg="white",
                activebackground=C["sidebar_hl"], activeforeground="white",
                relief="flat", anchor="w", padx=16, pady=10, cursor="hand2",
                command=lambda p=page: self._show_page(p))
            btn.pack(fill="x")
            self.nav_btns[page] = btn
        tk.Label(self.sidebar, text="v3.0 • 2025", font=FONT_SMALL,
                 bg=C["sidebar"], fg="#9FC5E8").pack(side="bottom", pady=12)

    def _build_statusbar(self):
        sb = tk.Frame(self.content, bg=C["border"], height=28)
        sb.pack(side="bottom", fill="x")
        tk.Label(sb, textvariable=self.status_var, font=FONT_SMALL,
                 bg=C["border"], fg=C["text"], padx=12).pack(side="left")

    def _show_page(self, page):
        for p, btn in self.nav_btns.items():
            btn.configure(bg=C["sidebar_hl"] if p==page else C["sidebar"])
        for w in self.content.winfo_children():
            if not (isinstance(w,tk.Frame) and w.cget("height")==28):
                w.destroy()
        {"home":self._page_home,"import":self._page_import,
         "update":self._page_update,"results":self._page_results,
         "export":self._page_export,
         "refs":self._page_refs,"log":self._page_log
        }.get(page, self._page_home)()

    def _scrollable(self, parent):
        canvas = tk.Canvas(parent, bg=C["bg"], highlightthickness=0)
        vsb    = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        frame  = tk.Frame(canvas, bg=C["bg"])
        frame.bind("<Configure>",
                   lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=frame, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(int(-1*(e.delta/120)),"units"))
        return frame

    # ══════════════════════════════════════════════════════════════
    # ANA SAYFA
    # ══════════════════════════════════════════════════════════════
    def _draw_home_kpi(self, parent):
        """Ana sayfadaki KPI grafik kartları."""
        try:
            import matplotlib; matplotlib.use("TkAgg")
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            import sys, os
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            from chart_style import make_kpi

            all_data = db_get_scope_totals()
            if not all_data: return
            years_data = filter_display_years(all_data)
            years = sorted(years_data.keys())
            if not years: return
            k1v  = [years_data[y]["k1"]    for y in years]
            k2v  = [years_data[y]["k2"]    for y in years]
            k3v  = [years_data[y]["k3"]    for y in years]
            totv = [years_data[y]["total"] for y in years]

            fig = make_kpi(years, k1v, k2v, k3v, totv)
            canvas = FigureCanvasTkAgg(fig, master=parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="x")

            # KPI vars'ı da güncelle (summary için)
            if hasattr(self, 'kpi_vars'):
                latest = years[-1]; d = years_data[latest]
                self.kpi_vars.get('toplam', tk.StringVar()).set(f"{d['total']:.1f}")
                self.kpi_vars.get('kapsam1',tk.StringVar()).set(f"{d['k1']:.1f}")
                self.kpi_vars.get('kapsam2',tk.StringVar()).set(f"{d['k2']:.1f}")
                self.kpi_vars.get('kapsam3',tk.StringVar()).set(f"{d['k3']:.1f}")
        except Exception as e:
            pass  # Grafik yoksa sessizce geç

    def _page_home(self):
        outer = tk.Frame(self.content, bg=C["bg"])
        outer.pack(fill="both", expand=True)
        frame = self._scrollable(outer)
        pad = tk.Frame(frame, bg=C["bg"])
        pad.pack(fill="both", expand=True, padx=24, pady=20)

        tk.Label(pad, text="UNSPED Karbon Ayak İzi", font=FONT_TITLE,
                 bg=C["bg"], fg=C["text"]).pack(anchor="w")
        tk.Label(pad, text="GHG Protocol tabanlı raporlama sistemi",
                 font=FONT, bg=C["bg"], fg=C["muted"]).pack(anchor="w", pady=(0,16))

        # KPI grafik — chart_style.py
        self.kpi_vars = {}
        kpi_canvas_frame = tk.Frame(pad, bg=C["bg"])
        kpi_canvas_frame.pack(fill="x", pady=(0,12))
        self._home_kpi_frame = kpi_canvas_frame
        self._draw_home_kpi(kpi_canvas_frame)

        # Hızlı işlemler
        section_label(pad, "Hızlı İşlemler")
        bf = tk.Frame(pad, bg=C["bg"]); bf.pack(fill="x")
        for text, color, cmd in [
            ("📥  Excel Seç & İmport", C["blue"],  lambda: self._show_page("import")),
            ("🔄  Faktörleri Güncelle",C["green"], lambda: self._show_page("update")),
            ("📊  Sonuçları Gör",      C["amber"], lambda: self._show_page("results")),
            ("📚  Referans Tablolar",  "#6B5B95",  lambda: self._show_page("refs")),
        ]:
            tk.Button(bf, text=text, font=FONT_BOLD, bg=color, fg="white",
                      relief="flat", padx=16, pady=8, cursor="hand2",
                      activebackground=color, command=cmd
                      ).pack(side="left", padx=(0,10))

        # DB Özeti
        section_label(pad, "Veritabanı Özeti")
        self.db_text = tk.Text(pad, height=10, font=FONT_MONO,
                               bg=C["card"], fg=C["text"], relief="flat",
                               state="disabled",
                               highlightbackground=C["border"], highlightthickness=1)
        self.db_text.pack(fill="x")
        self._load_db_summary()

    def _load_db_summary(self):
        try:
            from db.connection import SessionLocal
            from db.models import Company, ReportingPeriod, EmissionFactor, AuditLog
            with SessionLocal() as s:
                company   = s.query(Company).filter_by(name="UNSPED").first()
                periods   = s.query(ReportingPeriod).order_by(ReportingPeriod.year).all()
                ef_count  = s.query(EmissionFactor).count()
                log_count = s.query(AuditLog).count()
                last_imp  = s.query(AuditLog).filter_by(
                    action="excel_import").order_by(AuditLog.id.desc()).first()
            lines = []
            if company: lines.append(f"  Sirket     : {company.name}")
            lines.append(f"  Donemler   : {[p.year for p in periods]}")
            lines.append(f"  EF Kayitlar: {ef_count} adet")
            lines.append(f"  Toplam Log : {log_count} islem")
            if last_imp:
                lines.append(f"  Son Import : {str(last_imp.created_at)[:19]}")
            # Gerçek veriler
            all_data = db_get_scope_totals()
            display  = filter_display_years(all_data)
            if display:
                latest = max(display.keys())
                d = display[latest]
                lines.append(f"\n  [{latest}] K1={d['k1']:.1f}  K2={d['k2']:.1f}  K3={d['k3']:.1f}")
                lines.append(f"  [{latest}] TOPLAM: {d['total']:.2f} ton CO2e")
                if hasattr(self, 'kpi_vars'):
                    self.kpi_vars['toplam'].set(f"{d['total']:.1f}")
                    self.kpi_vars['kapsam1'].set(f"{d['k1']:.1f}")
                    self.kpi_vars['kapsam2'].set(f"{d['k2']:.1f}")
                    self.kpi_vars['kapsam3'].set(f"{d['k3']:.1f}")
            if hasattr(self,'db_text'):
                self.db_text.configure(state="normal")
                self.db_text.delete("1.0","end")
                self.db_text.insert("end", "\n".join(lines))
                self.db_text.configure(state="disabled")
        except Exception as e:
            if hasattr(self,'db_text'):
                self.db_text.configure(state="normal")
                self.db_text.delete("1.0","end")
                self.db_text.insert("end",
                    f"  DB baglantisi kurulamadi:\n  {e}\n\n  Cozum: python main.py calistirin")
                self.db_text.configure(state="disabled")

    # ══════════════════════════════════════════════════════════════
    # IMPORT SAYFASI
    # ══════════════════════════════════════════════════════════════
    def _page_import(self):
        frame = tk.Frame(self.content, bg=C["bg"])
        frame.pack(fill="both", expand=True, padx=24, pady=20)
        tk.Label(frame, text="Veri İmport", font=FONT_TITLE,
                 bg=C["bg"], fg=C["text"]).pack(anchor="w")
        tk.Label(frame, text="Excel dosyanızı seçin ve sisteme aktarın",
                 font=FONT, bg=C["bg"], fg=C["muted"]).pack(anchor="w", pady=(0,16))

        fi = card(frame, "Excel Dosyası")
        row = tk.Frame(fi, bg=C["card"]); row.pack(fill="x")
        tk.Entry(row, textvariable=self.excel_path, font=FONT,
                 width=55, relief="solid", bd=1).pack(side="left", padx=(0,8))
        tk.Button(row, text="Gözat...", font=FONT, bg=C["blue"], fg="white",
                  relief="flat", padx=12, pady=4, cursor="hand2",
                  command=self._browse_excel).pack(side="left")

        oi = card(frame, "Seçenekler")
        self.opt_update = tk.BooleanVar(value=False)
        tk.Checkbutton(oi, text="İmport öncesi emisyon faktörlerini güncelle",
                       variable=self.opt_update, font=FONT,
                       bg=C["card"], fg=C["text"], activebackground=C["card"]
                       ).pack(anchor="w")

        bf = tk.Frame(frame, bg=C["bg"]); bf.pack(fill="x", pady=8)
        self.import_btn = tk.Button(bf, text="▶  İmport Et",
                                    font=FONT_BOLD, bg=C["blue"], fg="white",
                                    relief="flat", padx=24, pady=10, cursor="hand2",
                                    command=self._run_import)
        self.import_btn.pack(side="left")
        tk.Button(bf, text="📊 Sonuçlara Git", font=FONT, bg=C["bg"],
                  fg=C["blue"], relief="flat", padx=12, pady=10, cursor="hand2",
                  command=lambda: self._show_page("results")
                  ).pack(side="left", padx=8)

        section_label(frame, "İmport Logu")
        self.import_log = scrolledtext.ScrolledText(
            frame, height=18, font=FONT_MONO,
            bg="#1E1E1E", fg="#D4D4D4", relief="flat",
            highlightbackground=C["border"], highlightthickness=1)
        self.import_log.pack(fill="both", expand=True)
        self.import_log.insert("end", "Import logu burada gorunecek...\n")

    def _browse_excel(self):
        p = filedialog.askopenfilename(
            title="Excel Seç",
            filetypes=[("Excel","*.xlsx *.xls"),("Tüm","*.*")])
        if p: self.excel_path.set(p)

    def _run_import(self):
        path = self.excel_path.get().strip()
        if not path:
            messagebox.showwarning("Uyarı","Lütfen Excel dosyası seçin."); return
        if not os.path.exists(path):
            messagebox.showerror("Hata",f"Dosya bulunamadı:\n{path}"); return
        self.import_btn.configure(state="disabled", text="⏳  Çalışıyor...")
        self.import_log.delete("1.0","end")
        self.status_var.set("İmport çalışıyor...")
        def run():
            old = sys.stdout
            try:
                sys.stdout = LogWriter(self.import_log)
                if self.opt_update.get():
                    from run_update import run_update
                    run_update(excel_path=path)
                from pipeline.Importer import run_import
                run_import(path)
                sys.stdout = old
                self.after(0, lambda: self._import_done(True))
            except Exception as e:
                import traceback; err = traceback.format_exc()
                sys.stdout = old
                self.after(0, lambda: self.import_log.insert("end",f"\nHATA:\n{err}"))
                self.after(0, lambda: self._import_done(False))
        threading.Thread(target=run, daemon=True).start()

    def _import_done(self, ok):
        self.import_btn.configure(state="normal", text="▶  İmport Et")
        self.status_var.set("✅ Tamamlandı" if ok else "❌ Hata")
        self._load_db_summary()
        if ok and messagebox.askyesno("Tamamlandı","İmport başarılı!\nSonuçları görüntüle?"):
            self._show_page("results")

    # ══════════════════════════════════════════════════════════════
    # FAKTÖR GÜNCELLEME
    # ══════════════════════════════════════════════════════════════
    def _page_update(self):
        frame = tk.Frame(self.content, bg=C["bg"])
        frame.pack(fill="both", expand=True, padx=24, pady=20)
        tk.Label(frame, text="Emisyon Faktörü Güncelleme", font=FONT_TITLE,
                 bg=C["bg"], fg=C["text"]).pack(anchor="w")
        tk.Label(frame, text="IPCC, DEFRA ve TC Enerji Bakanlığı EVÇED kaynaklarından günceller",
                 font=FONT, bg=C["bg"], fg=C["muted"]).pack(anchor="w", pady=(0,16))

        si = card(frame, "Güncellenecek Kaynaklar")
        self.src_vars = {}
        for key, label in [
            ("ipcc", "🔍  IPCC AR7 Kontrol (yeni rapor var mı?)"),
            ("teias","⚡  TEİAŞ/EVÇED Kontrol (yeni yıl var mı?)"),
            ("defra","🚛  Taşımacılık Faktörleri DEFRA (online indir)"),
        ]:
            v = tk.BooleanVar(value=True); self.src_vars[key] = v
            tk.Checkbutton(si, text=label, variable=v, font=FONT,
                           bg=C["card"], fg=C["text"], activebackground=C["card"]
                           ).pack(anchor="w", pady=2)

        ni = card(frame)
        tk.Label(ni, text="Not: GWP ve yakıt faktörleri Excel yeşil hücreler üzerinden yönetilir. "
                          "Excel seçilirse boş hücreler otomatik doldurulur.",
                 font=FONT_SMALL, bg=C["card"], fg=C["muted"],
                 wraplength=800, justify="left").pack(anchor="w")

        xi = card(frame, "Excel (opsiyonel — boş yeşil hücreler doldurulur)")
        xr = tk.Frame(xi, bg=C["card"]); xr.pack(fill="x")
        tk.Entry(xr, textvariable=self.excel_path, font=FONT,
                 width=50, relief="solid", bd=1).pack(side="left", padx=(0,8))
        tk.Button(xr, text="Gözat...", font=FONT, bg=C["blue"], fg="white",
                  relief="flat", padx=10, pady=3, cursor="hand2",
                  command=self._browse_excel).pack(side="left")

        ei = card(frame, "TC Enerji Bakanlığı EVÇED — Elektrik Grid Faktörleri")
        tk.Label(ei, text=f"Kaynak: {EVCED_URL}",
                 font=FONT_SMALL, bg=C["card"], fg=C["muted"]).pack(anchor="w", pady=(0,6))
        t = make_treeview(ei, ["Yıl","Üretim EF","Tüketim EF","Kaynak"],
                          [70,150,150,280], height=5)
        for y, d in sorted(EVCED_FACTORS.items()):
            t.insert("","end", values=(y, d["uretim"], d["tuketim"], d["kaynak"]))

        self.update_btn = tk.Button(frame, text="🔄  Güncellemeyi Başlat",
                                    font=FONT_BOLD, bg=C["green"], fg="white",
                                    relief="flat", padx=24, pady=10, cursor="hand2",
                                    command=self._run_update)
        self.update_btn.pack(anchor="w", pady=8)

        section_label(frame, "Güncelleme Logu")
        self.update_log = scrolledtext.ScrolledText(
            frame, height=10, font=FONT_MONO, bg="#1E1E1E", fg="#D4D4D4",
            relief="flat", highlightbackground=C["border"], highlightthickness=1)
        self.update_log.pack(fill="both", expand=True)
        self.update_log.insert("end", "Guncelleme logu burada gorunecek...\n")

    def _run_update(self):
        sources = [k for k,v in self.src_vars.items() if v.get()]
        if not sources:
            messagebox.showwarning("Uyarı","En az bir kaynak seçin."); return
        excel = self.excel_path.get().strip() or None
        self.update_btn.configure(state="disabled", text="⏳  Güncelleniyor...")
        self.update_log.delete("1.0","end")
        self.status_var.set("Güncelleniyor...")
        def run():
            old = sys.stdout
            try:
                sys.stdout = LogWriter(self.update_log)
                from run_update import run_update
                run_update(sources=sources, excel_path=excel)
                sys.stdout = old
                self.after(0, lambda: self._update_done(True))
            except Exception as e:
                import traceback; err = traceback.format_exc()
                sys.stdout = old
                self.after(0, lambda: self.update_log.insert("end",f"\nHATA:\n{err}"))
                self.after(0, lambda: self._update_done(False))
        threading.Thread(target=run, daemon=True).start()

    def _update_done(self, ok):
        self.update_btn.configure(state="normal", text="🔄  Güncellemeyi Başlat")
        self.status_var.set("✅ Tamamlandı" if ok else "❌ Hata")

    # ══════════════════════════════════════════════════════════════
    # SONUÇLAR & GRAFİK
    # ══════════════════════════════════════════════════════════════
    def _page_results(self):
        frame = tk.Frame(self.content, bg=C["bg"])
        frame.pack(fill="both", expand=True, padx=24, pady=20)
        tk.Label(frame, text="Sonuçlar & Grafikler", font=FONT_TITLE,
                 bg=C["bg"], fg=C["text"]).pack(anchor="w")
        tk.Label(frame, text="Emisyon hesaplama sonuçları", font=FONT,
                 bg=C["bg"], fg=C["muted"]).pack(anchor="w", pady=(0,8))

        tabf = tk.Frame(frame, bg=C["bg"]); tabf.pack(fill="x", pady=(0,8))
        self.res_tab = tk.StringVar(value="chart")
        for text, val in [("📊 Yıllık Grafik","chart"),
                          ("📈 Değişim Trendi","trend"),
                          ("📋 Kapsam Tablosu","table"),
                          ("🎯 Etki Analizi","impact"),
                          ("🏢 Kapsam 3 Detay","s3detail"),
                          ("👤 Kişi Başına","percap")]:
            tk.Radiobutton(tabf, text=text, variable=self.res_tab, value=val,
                           font=FONT, bg=C["bg"], fg=C["text"],
                           activebackground=C["bg"], cursor="hand2",
                           command=self._refresh_results
                           ).pack(side="left", padx=(0,12))

        self.res_content = tk.Frame(frame, bg=C["bg"])
        self.res_content.pack(fill="both", expand=True)
        self._refresh_results()

    def _refresh_results(self):
        for w in self.res_content.winfo_children(): w.destroy()
        {"chart":self._res_chart,"trend":self._res_trend,
         "table":self._res_table,"impact":self._res_impact,
         "s3detail":self._res_s3detail,"percap":self._res_percap
        }.get(self.res_tab.get(), self._res_chart)()

    def _res_chart(self):
        """Grouped Bar + Donut — chart_style.py stilinde."""
        try:
            import matplotlib; matplotlib.use("TkAgg")
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
            import sys, os
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            from chart_style import make_chart1, PAL

            all_data   = db_get_scope_totals()
            if not all_data: no_data_label(self.res_content); return
            years_data = filter_display_years(all_data)
            years = sorted(years_data.keys())
            k1v   = [years_data[y]["k1"]    for y in years]
            k2v   = [years_data[y]["k2"]    for y in years]
            k3v   = [years_data[y]["k3"]    for y in years]
            totv  = [years_data[y]["total"] for y in years]

            fig = make_chart1(years, k1v, k2v, k3v, totv)
            canvas = FigureCanvasTkAgg(fig, master=self.res_content)
            canvas.draw()
            # Toolbar — zoom, pan, kaydet
            toolbar_frame = tk.Frame(self.res_content, bg=C["bg"])
            toolbar_frame.pack(fill="x")
            toolbar = NavigationToolbar2Tk(canvas, toolbar_frame)
            toolbar.update()
            canvas.get_tk_widget().pack(fill="both", expand=True)
        except ImportError:
            no_data_label(self.res_content, "pip install matplotlib")
        except Exception as e:
            no_data_label(self.res_content, f"Grafik hatası: {e}")

    def _res_trend(self):
        """Stacked Bar + Toplam Trend — interaktif."""
        try:
            import matplotlib; matplotlib.use("TkAgg")
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
            import sys, os
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            from chart_style import make_chart2, PAL

            all_data   = db_get_scope_totals()
            if not all_data: no_data_label(self.res_content); return
            years_data = filter_display_years(all_data)
            years = sorted(years_data.keys())
            k1v   = [years_data[y]["k1"]    for y in years]
            k2v   = [years_data[y]["k2"]    for y in years]
            k3v   = [years_data[y]["k3"]    for y in years]
            totv  = [years_data[y]["total"] for y in years]

            fig = make_chart2(years, k1v, k2v, k3v, totv)
            canvas = FigureCanvasTkAgg(fig, master=self.res_content)
            canvas.draw()
            toolbar_frame = tk.Frame(self.res_content, bg=C["bg"])
            toolbar_frame.pack(fill="x")
            toolbar = NavigationToolbar2Tk(canvas, toolbar_frame)
            toolbar.update()
            canvas.get_tk_widget().pack(fill="both", expand=True)
        except ImportError:
            no_data_label(self.res_content, "pip install matplotlib")
        except Exception as e:
            no_data_label(self.res_content, f"Grafik hatası: {e}")

    def _res_table(self):
        """Kapsam Tablosu + KPI grafik."""
        all_data   = db_get_scope_totals()
        years_data = filter_display_years(all_data)
        if not years_data: no_data_label(self.res_content); return

        # KPI grafik üstte
        try:
            import matplotlib; matplotlib.use("TkAgg")
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            import sys, os
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            from chart_style import make_kpi

            years = sorted(years_data.keys())
            k1v   = [years_data[y]["k1"]    for y in years]
            k2v   = [years_data[y]["k2"]    for y in years]
            k3v   = [years_data[y]["k3"]    for y in years]
            totv  = [years_data[y]["total"] for y in years]

            fig_kpi = make_kpi(years, k1v, k2v, k3v, totv)
            c_kpi   = FigureCanvasTkAgg(fig_kpi, master=self.res_content)
            c_kpi.draw()
            c_kpi.get_tk_widget().pack(fill="x", pady=(0,4))
        except Exception:
            pass

        # Detay tablosu
        cols = ["Yıl","Kapsam 1","Kapsam 2","Kapsam 3","Toplam"]
        tree = make_treeview(self.res_content, cols,
                             [80,160,160,160,160], height=10)
        tree.tag_configure("total",   background=C["warning"], font=FONT_BOLD)
        tree.tag_configure("detail",  background=C["bg"])

        cat_labels = {
            "1.1":"  1.1 Sabit Yanma","1.2":"  1.2 Hareketli Yakma",
            "1.4":"  1.4 Soğutucu Gaz","2.1":"  2.1 Elektrik",
            "3.3":"  3.3 Personel Ulaşım","3.5":"  3.5 Uçak/Konaklama",
            "4.1":"  4.1 Satın Alınan","4.2":"  4.2 Sermaye",
            "4.3":"  4.3 Atık","4.5":"  4.5 Hizmetler","6.1":"  6.1 T&D",
        }
        for y in sorted(years_data.keys()):
            d = years_data[y]; cats = d.get("cats",{})
            tree.insert("","end",
                values=(y, f"{d['k1']:.2f}", f"{d['k2']:.2f}",
                        f"{d['k3']:.2f}", f"{d['total']:.2f}"),
                tags=("total",))
            for k, label in cat_labels.items():
                v = cats.get(k, 0)
                if v > 0:
                    tree.insert("","end",
                        values=(label,"","","",f"{v:.4f}"),
                        tags=("detail",))

    def _res_impact(self):
        """
        Etki Analizi — SADECE Kapsam 3 (3.1 → 6.1).
        Tüm kaynaklar listelenir, %95 eşiği işaretlenir.
        """
        all_data   = db_get_scope_totals()
        years_data = filter_display_years(all_data)
        if not years_data: no_data_label(self.res_content); return

        latest = max(years_data.keys())
        d      = years_data[latest]
        cats   = d.get("cats", {})

        # Başlık
        hf = tk.Frame(self.res_content, bg=C["bg"]); hf.pack(fill="x", pady=(0,8))
        tk.Label(hf,
                 text=f"Etki Analizi — {latest} | Kapsam 3 Kaynak Analizi (%95 Eşiği)",
                 font=FONT_BOLD, bg=C["bg"], fg=C["text"]).pack(side="left")
        s3_total = d["k3"]
        tk.Label(hf, text=f"  Kapsam 3 Toplam: {s3_total:.2f} ton CO2e",
                 font=FONT, bg=C["bg"], fg=C["muted"]).pack(side="left")

        # Sadece Kapsam 3 kategorileri
        s3_cats = [
            ("3.1","Hammadde Sevkiyatı"),
            ("3.2","Ürün Sevkiyatı"),
            ("3.3","Personel Ulaşım"),
            ("3.4","İş Seyahati - Karayolu"),
            ("3.5","Uçak & Konaklama"),
            ("4.1","Satın Alınan Malzemeler"),
            ("4.2","Sermaye Varlıkları"),
            ("4.3","Atık Bertarafı"),
            ("4.4","Kiralanan Ekipmanlar"),
            ("4.5","Hizmet Alımları"),
            ("6.1","T&D Kayıpları"),
        ]

        sources = sorted(
            [("Kapsam 3", cat, label, cats.get(cat,0))
             for cat, label in s3_cats],
            key=lambda x: x[3], reverse=True
        )

        if s3_total <= 0:
            no_data_label(self.res_content, "Kapsam 3 verisi yok.")
            return

        cols = ["Sıra","Kapsam","Kategori","Emisyon Kaynağı",
                "CO2e (ton)","Pay (%)","Kümülatif (%)","Durum"]
        tree = make_treeview(self.res_content, cols,
                             [45,80,70,190,110,70,95,80], height=14)
        tree.tag_configure("above", background=C["impact"])   # turuncu — %95 içi
        tree.tag_configure("below", background=C["bg"])        # normal — %95 dışı
        tree.tag_configure("thresh", background=C["red"],
                           foreground="white")                  # kırmızı — eşik

        cumulative = 0
        threshold_added = False
        sira = 1

        for scope, cat, label, val in sources:
            pct = round(val / s3_total * 100, 2) if s3_total > 0 else 0
            cumulative = round(cumulative + pct, 2)

            # %95 eşiği geçildi mi?
            if cumulative > 95 and not threshold_added:
                tree.insert("","end",
                    values=("","","",
                            f"▲ %95 EŞİĞİ — Aşağıdaki kaynaklar K3'ün %5'inden az",
                            "","","",""),
                    tags=("thresh",))
                threshold_added = True

            tag = "above" if cumulative <= 95 else "below"
            durum = "✅ Kritik" if cumulative <= 95 else "—"
            tree.insert("","end", values=(
                sira, scope, cat, label,
                f"{val:.2f}", f"{pct:.1f}%", f"{cumulative:.1f}%", durum
            ), tags=(tag,))
            sira += 1

        # Açıklama
        tk.Label(self.res_content,
                 text="🟠 Turuncu = K3 emisyonlarının %95'ini oluşturan kritik kaynaklar",
                 font=FONT_SMALL, bg=C["bg"], fg=C["muted"]).pack(anchor="w", pady=4)

    def _res_s3detail(self):
        """Kapsam 3 alt kategoriler — açılır/kapanır gruplar."""
        outer = tk.Frame(self.res_content, bg=C["bg"])
        outer.pack(fill="both", expand=True)
        sf = self._scrollable(outer)
        pad = tk.Frame(sf, bg=C["bg"]); pad.pack(fill="both", expand=True, padx=8, pady=8)

        tk.Label(pad, text="Kapsam 3 Alt Kategori Detayı",
                 font=FONT_BOLD, bg=C["bg"], fg=C["text"]).pack(anchor="w", pady=(0,8))
        tk.Label(pad, text="Gruba tıklayarak açın/kapatın.",
                 font=FONT_SMALL, bg=C["bg"], fg=C["muted"]).pack(anchor="w", pady=(0,12))

        all_data   = db_get_scope_totals()
        years_data = filter_display_years(all_data)
        if not years_data: no_data_label(pad,"Veri yok."); return

        latest = max(years_data.keys())
        cats   = years_data[latest].get("cats", {})

        groups = {
            "3 — Kapsam 3 (Dolaylı)": {
                "color": C["s3"],
                "children": {
                    "3.1 — Hammadde Sevkiyatı":  cats.get("3.1",0),
                    "3.2 — Ürün Sevkiyatı":      cats.get("3.2",0),
                    "3.3 — Personel Ulaşım":     cats.get("3.3",0),
                    "3.4 — İş Seyahati":         cats.get("3.4",0),
                    "3.5 — Uçak & Konaklama":    cats.get("3.5",0),
                }
            },
            "4 — Tedarik Zinciri": {
                "color": "#8B4513",
                "children": {
                    "4.1 — Satın Alınan Mallar": cats.get("4.1",0),
                    "4.2 — Sermaye Varlıkları":  cats.get("4.2",0),
                    "4.3 — Atık Bertarafı":      cats.get("4.3",0),
                    "4.4 — Kiralanan Ekipman":   cats.get("4.4",0),
                    "4.5 — Hizmet Alımları":     cats.get("4.5",0),
                }
            },
            "6 — Elektrik T&D": {
                "color": "#4B0082",
                "children": {
                    "6.1 — T&D Kayıpları": cats.get("6.1",0),
                }
            },
        }

        for group_name, gdata in groups.items():
            color    = gdata["color"]
            children = gdata["children"]
            g_total  = sum(children.values())
            state_var = tk.BooleanVar(value=True)

            hf = tk.Frame(pad, bg=color); hf.pack(fill="x", pady=(4,0))
            cf = tk.Frame(pad, bg=C["card"],
                          highlightbackground=color, highlightthickness=1)

            def toggle(sv=state_var, c=cf):
                c.pack(fill="x") if sv.get() else c.pack_forget()

            tk.Button(hf,
                text=f"  ▼  {group_name}    {g_total:.2f} ton CO2e",
                font=FONT_BOLD, bg=color, fg="white",
                relief="flat", anchor="w", padx=12, pady=8, cursor="hand2",
                activebackground=color,
                command=lambda sv=state_var, c=cf: (sv.set(not sv.get()), toggle(sv, c))
            ).pack(fill="x")

            for i, (cat_name, val) in enumerate(children.items()):
                row_bg = C["card"] if i%2==0 else "#F8F8F5"
                rf = tk.Frame(cf, bg=row_bg); rf.pack(fill="x")
                tk.Label(rf, text=f"    {cat_name}", font=FONT,
                         bg=row_bg, fg=C["text"], padx=16, pady=6,
                         anchor="w").pack(side="left", fill="x", expand=True)
                pct = (val/g_total*100) if g_total else 0
                tk.Label(rf, text=f"{pct:.1f}%",
                         font=FONT_MONO, bg=row_bg, fg=C["muted"],
                         padx=8).pack(side="right")
                tk.Label(rf, text=f"{val:.4f} ton CO2e",
                         font=FONT_MONO, bg=row_bg, fg=C["text"],
                         padx=16).pack(side="right")
            cf.pack(fill="x")

    def _res_percap(self):
        """Kişi başına emisyon — PERSONNEL sayfasından + scope_totals."""
        outer = tk.Frame(self.res_content, bg=C["bg"])
        outer.pack(fill="both", expand=True)
        sf = self._scrollable(outer)
        pad = tk.Frame(sf, bg=C["bg"]); pad.pack(fill="both", expand=True, padx=8, pady=8)

        tk.Label(pad, text="Kişi Başına Emisyon — Lokasyon Bazında",
                 font=FONT_BOLD, bg=C["bg"], fg=C["text"]).pack(anchor="w", pady=(0,4))
        tk.Label(pad, text="Veriler PERSONNEL sayfasından + DB'den otomatik hesaplanır.",
                 font=FONT_SMALL, bg=C["bg"], fg=C["muted"]).pack(anchor="w", pady=(0,12))

        excel = self.excel_path.get().strip()
        if not excel or not os.path.exists(excel):
            no_data_label(pad, "Excel dosyasını 'Veri İmport' sayfasından seçin."); return

        try:
            from openpyxl import load_workbook
            wb  = load_workbook(excel, read_only=True, data_only=True)
            ws  = wb['PERSONNEL']
            # Personel verisi
            personnel_by_year = {}
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row[0] or not row[1] or not row[2]: continue
                try:
                    year = int(float(str(row[0])))
                    if not (2000 <= year <= 2100): continue
                    personnel_by_year.setdefault(year,[]).append({
                        "location": str(row[1]).strip(),
                        "headcount": int(float(str(row[2])))
                    })
                except: continue

            if not personnel_by_year:
                no_data_label(pad, "PERSONNEL sayfasında veri bulunamadı."); return

            # Emisyon verisi
            all_data = db_get_scope_totals()
            if not all_data:
                no_data_label(pad, "DB'de emisyon verisi yok."); return

            cols = ["Yıl","Lokasyon","Personel",
                    "K1 CO2e","K2 CO2e","K3 CO2e",
                    "Toplam CO2e","ton CO2e / Kişi"]
            tree = make_treeview(pad, cols,
                                 [60,180,80,90,90,90,110,120], height=16)
            tree.tag_configure("total",  background=C["warning"], font=FONT_BOLD)
            tree.tag_configure("normal", background=C["bg"])

            for year in sorted(personnel_by_year.keys()):
                if str(year) not in all_data: continue
                personnel  = personnel_by_year[year]
                d          = all_data[str(year)]
                s1,s2,s3   = d["k1"], d["k2"], d["k3"]
                grand      = d["total"]
                total_head = sum(p["headcount"] for p in personnel)
                if total_head == 0: continue

                for p in personnel:
                    ratio     = p["headcount"] / total_head
                    loc_s1    = round(s1*ratio, 4)
                    loc_s2    = round(s2*ratio, 4)
                    loc_s3    = round(s3*ratio, 4)
                    loc_total = round(loc_s1+loc_s2+loc_s3, 4)
                    per_cap   = round(loc_total/p["headcount"],6) if p["headcount"] else 0
                    tree.insert("","end", values=(
                        year, p["location"], p["headcount"],
                        loc_s1, loc_s2, loc_s3, loc_total, per_cap
                    ), tags=("normal",))

                # Yıl toplamı
                per_cap_total = round(grand/total_head,6) if total_head else 0
                tree.insert("","end", values=(
                    year, f"ŞİRKET TOPLAMI ({year})", total_head,
                    round(s1,4), round(s2,4), round(s3,4),
                    round(grand,4), per_cap_total
                ), tags=("total",))

        except Exception as e:
            import traceback
            no_data_label(pad, f"Hata: {e}")

    # ══════════════════════════════════════════════════════════════
    # REFERANS TABLOLAR
    # ══════════════════════════════════════════════════════════════
    def _page_refs(self):
        frame = tk.Frame(self.content, bg=C["bg"])
        frame.pack(fill="both", expand=True)
        outer = tk.Frame(frame, bg=C["bg"]); outer.pack(fill="both", expand=True)
        sf  = self._scrollable(outer)
        pad = tk.Frame(sf, bg=C["bg"]); pad.pack(fill="both", expand=True, padx=24, pady=20)

        tk.Label(pad, text="Referans Tablolar", font=FONT_TITLE,
                 bg=C["bg"], fg=C["text"]).pack(anchor="w")
        tk.Label(pad, text="GWP, Emisyon Faktörleri, NCV ve EVÇED",
                 font=FONT, bg=C["bg"], fg=C["muted"]).pack(anchor="w", pady=(0,16))

        # 1. GWP
        section_label(pad, "1. GWP-100 Değerleri — IPCC AR6 2021")
        gi = card(pad, pady=16)
        tk.Label(gi, text="Kaynak: IPCC AR6 WGI Tablo 7.SM.7",
                 font=FONT_SMALL, bg=C["card"], fg=C["muted"]).pack(anchor="w", pady=(0,6))
        t = make_treeview(gi, ["Formül","Gaz Adı","GWP-100","Kaynak"],
                          [90,200,90,220], height=9)
        for row in GWP_TABLE: t.insert("","end",values=row)

        # 2. EF
        section_label(pad, "2. Yakıt Emisyon Faktörleri — IPCC 2006 GL")
        ei = card(pad, pady=16)
        tk.Label(ei, text="Kaynak: IPCC 2006 Guidelines Vol.2 Table 2.2",
                 font=FONT_SMALL, bg=C["card"], fg=C["muted"]).pack(anchor="w", pady=(0,6))
        t2 = make_treeview(ei,
             ["Yakıt","Birim","EF CO2","EF CH4","EF N2O","NCV","Yoğunluk","Kaynak"],
             [110,55,80,80,80,70,90,160], height=7)
        for row in EF_TABLE: t2.insert("","end",values=row)

        # 3. NCV
        section_label(pad, "3. Net Kalorifik Değer — IPCC 2006 GL")
        ni = card(pad, pady=16)
        tk.Label(ni, text="Kaynak: IPCC 2006 Guidelines Vol.2 Table 1.2",
                 font=FONT_SMALL, bg=C["card"], fg=C["muted"]).pack(anchor="w", pady=(0,6))
        t3 = make_treeview(ni,
             ["Yakıt","Birim","NCV","NCV Birimi","Yoğunluk","Yoğunluk Birimi","Kaynak"],
             [120,55,70,70,85,110,180], height=7)
        for row in NCV_TABLE: t3.insert("","end",values=row)

        # 4. EVÇED
        section_label(pad, "4. TC Enerji Bakanlığı EVÇED — Elektrik Emisyon Faktörleri")
        evi = card(pad, pady=16)
        tk.Label(evi, text=f"Kaynak: {EVCED_URL}",
                 font=FONT_SMALL, bg=C["card"], fg=C["blue"]).pack(anchor="w", pady=(0,6))
        ev_dates = {2019:"Mart 2021",2020:"Mart 2021",2021:"Eylül 2022",
                    2022:"Aralık 2024",2023:"Aralık 2025"}
        t4 = make_treeview(evi,
             ["Yıl","Üretim EF (tCO2/MWh)","Tüketim EF (tCO2/MWh)",
              "Yayın Tarihi","Kaynak"],
             [55,160,160,150,260], height=6)
        for y, d in sorted(EVCED_FACTORS.items()):
            t4.insert("","end", values=(y,d["uretim"],d["tuketim"],
                                        ev_dates.get(y,"—"),d["kaynak"]))
        tk.Label(evi,
                 text="Üretim EF = elektrik üretim noktası. "
                      "Tüketim EF = şebekeden alınan elektrik (T&D dahil).",
                 font=FONT_SMALL, bg=C["card"], fg=C["muted"],
                 wraplength=900, justify="left").pack(anchor="w", pady=(8,0))

    # ══════════════════════════════════════════════════════════════
    # WORD RAPOR EXPORT
    # ══════════════════════════════════════════════════════════════
    def _page_export(self):
        outer = tk.Frame(self.content, bg=C["bg"])
        outer.pack(fill="both", expand=True)
        frame = self._scrollable(outer)
        pad = tk.Frame(frame, bg=C["bg"])
        pad.pack(fill="both", expand=True, padx=24, pady=20)

        tk.Label(pad, text="Word Rapor Export", font=FONT_TITLE,
                 bg=C["bg"], fg=C["text"]).pack(anchor="w")
        tk.Label(pad, text="Seçtiğiniz yıllar ve bölümleri Word dosyasına aktarın",
                 font=FONT, bg=C["bg"], fg=C["muted"]).pack(anchor="w", pady=(0,16))

        # ── Yıl seçimi ────────────────────────────────────────────
        yi = card(pad, "Yılları Seçin")
        all_data = db_get_scope_totals()
        all_years = sorted(all_data.keys()) if all_data else []

        self.export_year_vars = {}
        if all_years:
            yf = tk.Frame(yi, bg=C["card"]); yf.pack(fill="x")
            for y in all_years:
                v = tk.BooleanVar(value=True)
                self.export_year_vars[y] = v
                tk.Checkbutton(yf, text=y, variable=v, font=FONT,
                               bg=C["card"], fg=C["text"],
                               activebackground=C["card"]
                               ).pack(side="left", padx=8)
        else:
            tk.Label(yi, text="Henüz import edilmiş veri yok.",
                     font=FONT, bg=C["card"], fg=C["muted"]).pack(anchor="w")

        # ── Bölüm seçimi ──────────────────────────────────────────
        si = card(pad, "Rapor Bölümleri")
        self.export_sec_vars = {}
        sections = [
            ("intro",           "📝  Giriş"),
            ("technical",       "⚙️   Bölüm 1: Teknik Bilgiler (metodoloji kısa metin)"),
            ("boundaries",      "🏢  Bölüm 2-3: Kuruluş ve Raporlama Sınırları"),
            ("scope1",          "🔵  Bölüm 4.1: Kapsam 1 — Doğrudan Emisyonlar"),
            ("scope2",          "🟢  Bölüm 4.2: Kapsam 2 — Elektrik"),
            ("scope3",          "🟠  Bölüm 4.3: Kapsam 3 — Dolaylı Emisyonlar"),
            ("activity",        "📋  Bölüm 4.3.1: Faaliyet Verileri (%95 eşiği)"),
            ("charts",          "📈  Bölüm 4.5: Grafikler (Bar + Pasta + Trend)"),
            ("summary",         "📊  Bölüm 4.4: Kapsamlara Göre Özet Tablo"),
            ("impact",          "🎯  Bölüm 4.6: Etki Analizi — %95 (metin)"),
            ("percap",          "👤  Bölüm 4.7: Kişi Başına Emisyon"),
            ("yearly",          "📅  Bölüm 5: Yıllara Göre GHG Verileri"),
            ("ghg_declaration", "📋  Bölüm 5: GHG Emisyon Beyanı (ISO 14064-1)"),
            ("conclusion",      "✅  Bölüm 6: Sonuç Tablosu"),
            ("recommendations", "💡  Bölüm 7: Öneriler"),
            ("methodology",     "📚  Bölüm 8-9: Metodoloji + NKD + GWP + EF Tabloları"),
        ]
        for key, label in sections:
            v = tk.BooleanVar(value=True)
            self.export_sec_vars[key] = v
            tk.Checkbutton(si, text=label, variable=v, font=FONT,
                           bg=C["card"], fg=C["text"],
                           activebackground=C["card"]
                           ).pack(anchor="w", pady=2)

        # ── Excel dosyası (kişi başına için) ──────────────────────
        ei = card(pad, "Excel Dosyası (Kişi Başına bölümü için gerekli)")
        er = tk.Frame(ei, bg=C["card"]); er.pack(fill="x")
        tk.Entry(er, textvariable=self.excel_path, font=FONT,
                 width=50, relief="solid", bd=1).pack(side="left", padx=(0,8))
        tk.Button(er, text="Gözat...", font=FONT, bg=C["blue"], fg="white",
                  relief="flat", padx=10, pady=3, cursor="hand2",
                  command=self._browse_excel).pack(side="left")

        # ── Export butonu ─────────────────────────────────────────
        bf = tk.Frame(pad, bg=C["bg"]); bf.pack(fill="x", pady=12)
        self.export_btn = tk.Button(bf,
            text="📄  Word Rapor Oluştur",
            font=FONT_BOLD, bg=C["blue"], fg="white",
            relief="flat", padx=24, pady=12, cursor="hand2",
            command=self._run_export)
        self.export_btn.pack(side="left")

        # ── Log ───────────────────────────────────────────────────
        section_label(pad, "Durum")
        self.export_log = tk.Text(pad, height=8, font=FONT_MONO,
                                  bg=C["card"], fg=C["text"], relief="flat",
                                  state="disabled",
                                  highlightbackground=C["border"], highlightthickness=1)
        self.export_log.pack(fill="x")
        self._export_log("Ayarları seçin ve 'Word Rapor Oluştur' butonuna tıklayın.")

    def _export_log(self, msg):
        try:
            self.export_log.configure(state="normal")
            self.export_log.insert("end", msg + "\n")
            self.export_log.see("end")
            self.export_log.configure(state="disabled")
        except: pass

    def _run_export(self):
        import json, subprocess, tempfile
        from datetime import datetime

        selected_years = [y for y, v in self.export_year_vars.items() if v.get()]
        selected_secs  = [k for k, v in self.export_sec_vars.items() if v.get()]

        if not selected_years:
            messagebox.showwarning("Uyarı","En az bir yıl seçin."); return
        if not selected_secs:
            messagebox.showwarning("Uyarı","En az bir bölüm seçin."); return

        self.export_btn.configure(state="disabled", text="⏳  Oluşturuluyor...")
        self.export_log.configure(state="normal")
        self.export_log.delete("1.0","end")
        self.export_log.configure(state="disabled")

        def run():
            try:
                self.after(0, lambda: self._export_log("Veriler DB'den okunuyor..."))

                all_data = db_get_scope_totals()
                if not all_data:
                    self.after(0, lambda: self._export_log("HATA: DB boş."))
                    self.after(0, lambda: self.export_btn.configure(
                        state="normal", text="📄  Word Rapor Oluştur"))
                    return

                # Seçili yılları filtrele
                scope_totals = {y: all_data[y] for y in selected_years if y in all_data}
                cats_data    = {y: all_data[y]["cats"] for y in selected_years if y in all_data}

                # Kişi başına veri (Excel'den)
                per_capita = []
                excel = self.excel_path.get().strip()
                if excel and os.path.exists(excel) and "percap" in selected_secs:
                    try:
                        from openpyxl import load_workbook
                        wb  = load_workbook(excel, read_only=True, data_only=True)
                        ws  = wb['PERSONNEL']
                        personnel_by_year = {}
                        for row in ws.iter_rows(min_row=4, values_only=True):
                            if not row[0] or not row[1] or not row[2]: continue
                            try:
                                year = str(int(float(str(row[0]))))
                                if year not in selected_years: continue
                                personnel_by_year.setdefault(year,[]).append({
                                    "location": str(row[1]).strip(),
                                    "headcount": int(float(str(row[2])))
                                })
                            except: continue

                        for year in selected_years:
                            if year not in personnel_by_year: continue
                            if year not in all_data: continue
                            personnel = personnel_by_year[year]
                            d = all_data[year]
                            total_head = sum(p["headcount"] for p in personnel)
                            if not total_head: continue
                            for p in personnel:
                                ratio = p["headcount"] / total_head
                                lt = round((d["k1"]+d["k2"]+d["k3"])*ratio, 4)
                                per_capita.append({
                                    "year": year,
                                    "location": p["location"],
                                    "headcount": p["headcount"],
                                    "k1": round(d["k1"]*ratio, 4),
                                    "k2": round(d["k2"]*ratio, 4),
                                    "k3": round(d["k3"]*ratio, 4),
                                    "total": lt,
                                    "per_cap": round(lt/p["headcount"],6) if p["headcount"] else 0,
                                    "is_total": False
                                })
                            per_capita.append({
                                "year": year,
                                "location": f"ŞİRKET TOPLAMI ({year})",
                                "headcount": total_head,
                                "k1": round(d["k1"],4), "k2": round(d["k2"],4),
                                "k3": round(d["k3"],4), "total": round(d["total"],4),
                                "per_cap": round(d["total"]/total_head,6),
                                "is_total": True
                            })
                    except Exception as pe:
                        self.after(0, lambda: self._export_log(
                            f"Uyarı: Kisi basina veri okunamadi: {pe}"))

                # Alt kategori detayları DB'den çek
                detail_data = {}
                try:
                    from db.connection import SessionLocal
                    from db.models import (StationaryCombustion, MobileCombustion,
                                          Refrigerant, ElectricityConsumption,
                                          FreightEmission, EmployeeCommuting,
                                          PurchasedGoods, CapitalGoods, ReportingPeriod)
                    with SessionLocal() as s:
                        for year in selected_years:
                            period = s.query(ReportingPeriod).filter_by(year=int(year)).first()
                            if not period: continue
                            pid = period.id
                            detail_data[year] = {
                                "stationary": [
                                    {"source": r.emission_source, "fuel": r.fuel_type,
                                     "value": r.activity_value, "unit": r.activity_unit,
                                     "co2e": round(r.co2e_total or 0, 4)}
                                    for r in s.query(StationaryCombustion).filter_by(period_id=pid).all()
                                ],
                                "mobile": [
                                    {"source": r.emission_source, "fuel": r.fuel_type,
                                     "value": r.activity_value, "unit": r.activity_unit,
                                     "co2e": round(r.co2e_total or 0, 4)}
                                    for r in s.query(MobileCombustion).filter_by(period_id=pid).all()
                                ],
                                "refrigerant": [
                                    {"source": r.gas_type, "fuel": r.gas_type,
                                     "value": r.activity_value, "unit": r.activity_unit,
                                     "co2e": round(r.co2e_total or 0, 4)}
                                    for r in s.query(Refrigerant).filter_by(period_id=pid).all()
                                ],
                            }
                except Exception as de:
                    detail_data = {}

                # JSON data dosyası
                report_data = {
                    "years": selected_years,
                    "scope_totals": scope_totals,
                    "cats": cats_data,
                    "detail": detail_data,
                    "per_capita": per_capita,
                    "company": "UNSPED",
                    "generated_at": datetime.now().strftime("%d.%m.%Y"),
                    "selected_sections": selected_secs
                }

                tmp_json = os.path.join(tempfile.gettempdir(), "unsped_report_data.json")
                with open(tmp_json, "w", encoding="utf-8") as f:
                    json.dump(report_data, f, ensure_ascii=False)

                self.after(0, lambda: self._export_log("Word raporu oluşturuluyor..."))

                # Python script yolu
                script_dir = os.path.dirname(os.path.abspath(__file__))
                py_script  = os.path.join(script_dir, "word_export", "generate_report.py")

                if not os.path.exists(py_script):
                    self.after(0, lambda: self._export_log(
                        "HATA: generate_report.py bulunamadi.\n"
                        "word_export klasorunu kontrol edin."))
                    self.after(0, lambda: self.export_btn.configure(
                        state="normal", text="📄  Word Rapor Oluştur"))
                    return

                # Kaydetme yeri seç
                out_path = filedialog.asksaveasfilename(
                    title="Raporu Kaydet",
                    defaultextension=".docx",
                    initialfile=f"UNSPED_Karbon_Raporu_{'-'.join(selected_years)}.docx",
                    filetypes=[("Word Belgesi","*.docx"),("Tüm","*.*")]
                )
                if not out_path:
                    self.after(0, lambda: self.export_btn.configure(
                        state="normal", text="📄  Word Rapor Oluştur"))
                    return

                # Python ile rapor oluştur (Node.js gerekmez)
                result = subprocess.run(
                    [sys.executable, py_script, tmp_json, out_path],
                    capture_output=True, text=True, timeout=120,
                    cwd=script_dir
                )

                if result.returncode == 0 and os.path.exists(out_path):
                    size_kb = os.path.getsize(out_path) // 1024
                    self.after(0, lambda: self._export_log(
                        f"Basarili! {out_path} ({size_kb} KB)"))
                    self.after(0, lambda p=out_path, sy=selected_years, ss=selected_secs:
                        messagebox.showinfo(
                        "Tamamlandı",
                        f"Word raporu olusturuldu!\n\n"
                        f"Yillar: {', '.join(sy)}\n"
                        f"Bolumler: {len(ss)} bolum\n"
                        f"Dosya: {os.path.basename(p)}"
                    ))
                    try:
                        import subprocess as sp
                        sp.Popen(["start", "", out_path], shell=True)
                    except: pass
                else:
                    err_out = (result.stderr or "") + (result.stdout or "")
                    err = err_out[:600] if err_out.strip() else "Bilinmeyen hata"
                    self.after(0, lambda e=err: self._export_log("HATA:\n" + str(e)))

            except Exception as e:
                import traceback
                self.after(0, lambda: self._export_log(f"HATA: {traceback.format_exc()[:300]}"))
            finally:
                self.after(0, lambda: self.export_btn.configure(
                    state="normal", text="📄  Word Rapor Oluştur"))

        threading.Thread(target=run, daemon=True).start()

    # ══════════════════════════════════════════════════════════════
    # LOG & REVİZYONLAR
    # ══════════════════════════════════════════════════════════════
    def _page_log(self):
        frame = tk.Frame(self.content, bg=C["bg"])
        frame.pack(fill="both", expand=True, padx=24, pady=20)
        tk.Label(frame, text="Log & Revizyonlar", font=FONT_TITLE,
                 bg=C["bg"], fg=C["text"]).pack(anchor="w")
        tk.Label(frame, text="Tüm import ve güncelleme revizyonları",
                 font=FONT, bg=C["bg"], fg=C["muted"]).pack(anchor="w", pady=(0,8))

        tf = tk.Frame(frame, bg=C["bg"]); tf.pack(fill="x", pady=(0,8))
        self.log_tab = tk.StringVar(value="revisions")
        for text, val in [("📋 Revizyonlar","revisions"),("🔧 Tüm İşlemler","all")]:
            tk.Radiobutton(tf, text=text, variable=self.log_tab, value=val,
                           font=FONT, bg=C["bg"], fg=C["text"],
                           activebackground=C["bg"], cursor="hand2",
                           command=self._refresh_log).pack(side="left", padx=(0,16))
        tk.Button(tf, text="🔄 Yenile", font=FONT, bg=C["blue"], fg="white",
                  relief="flat", padx=12, pady=4, cursor="hand2",
                  command=self._refresh_log).pack(side="right")
        self.log_content = tk.Frame(frame, bg=C["bg"])
        self.log_content.pack(fill="both", expand=True)
        self._refresh_log()

    def _refresh_log(self):
        for w in self.log_content.winfo_children(): w.destroy()
        if self.log_tab.get() == "revisions":
            self._log_revisions()
        else:
            self._log_all()

    def _log_revisions(self):
        revs = db_get_revisions()
        info = tk.Frame(self.log_content, bg=C["warning"],
                        highlightbackground=C["border"], highlightthickness=1)
        info.pack(fill="x", pady=(0,8))
        inf = tk.Frame(info, bg=C["warning"], padx=16, pady=8); inf.pack(fill="x")
        tk.Label(inf,
                 text="Her import aynı yıl için yeni revizyon oluşturur.\n"
                      "2025-R1 = ilk import, 2025-R2 = ikinci import",
                 font=FONT_SMALL, bg=C["warning"], fg=C["text"],
                 justify="left").pack(anchor="w")

        if not revs:
            no_data_label(self.log_content,"Henüz import yapılmadı."); return

        cols = ["Revizyon ID","Yıl","Tarih","Kapsam 1","Kapsam 2","Kapsam 3","Toplam","Durum"]
        tree = make_treeview(self.log_content, cols,
                             [100,60,150,110,110,110,120,80], height=16)

        latest_by_year = {}
        for r in revs:
            y = r["year"]
            if y not in latest_by_year: latest_by_year[y] = r["id"]

        for r in revs:
            is_latest = latest_by_year.get(r["year"]) == r["id"]
            tag = "latest" if is_latest else "old"
            tree.insert("","end", values=(
                r["id"], r["year"], r["ts"],
                f"{r['k1']:.2f}", f"{r['k2']:.2f}",
                f"{r['k3']:.2f}", f"{r['total']:.2f}",
                "✅ Güncel" if is_latest else "📦 Arşiv"
            ), tags=(tag,))

        tree.tag_configure("latest", background=C["success"], font=FONT_BOLD)
        tree.tag_configure("old",    background=C["bg"])
        tk.Label(self.log_content,
                 text="✅ Yeşil = geçerli revizyon   📦 Gri = arşiv",
                 font=FONT_SMALL, bg=C["bg"], fg=C["muted"]).pack(anchor="w", pady=4)

    def _log_all(self):
        try:
            from db.connection import SessionLocal
            from db.models import AuditLog
            cols = ["Tarih","İşlem","Kapsam","Durum","Notlar"]
            tree = make_treeview(self.log_content, cols,
                                 [150,140,100,80,450], height=18)
            with SessionLocal() as s:
                for log in s.query(AuditLog).order_by(AuditLog.id.desc()).limit(200).all():
                    tag = "success" if log.status=="success" else "fail"
                    tree.insert("","end", values=(
                        str(log.created_at)[:19] if log.created_at else "—",
                        log.action or "—", log.scope or "—",
                        log.status or "—", (log.notes or "")[:80]
                    ), tags=(tag,))
            tree.tag_configure("success", background=C["success"])
            tree.tag_configure("fail",    background=C["danger"])
        except Exception as e:
            no_data_label(self.log_content, f"DB Hatası: {e}")


# ══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = UNSPEDDashboard()
    app.mainloop()