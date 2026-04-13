"""
UNSPED Carbon Footprint — Faktör Güncelleme Sistemi
====================================================
Kullanım:
    python run_update.py --excel data/UNSPED_Karbon_Veri_Girisi_v3.xlsx
    python run_update.py --source defra --excel data/...xlsx
    python run_update.py --source gwp --excel data/...xlsx

Mimari:
    KATMAN 1 — Online Kontrol (sadece DEFRA gerçekten indirilir)
        DEFRA  : GOV.UK'dan yeni yayın var mı → indir → DB + Excel
        IPCC   : AR7 yayınlandı mı → sadece uyar, değer değişmez
        TEİAŞ  : Yeni yıl tespit edildi mi → sadece uyar

    KATMAN 2 — Excel Boş Hücre Doldurma
        GWP    : 1.4_Refrigerants — GWP sütunu boşsa GWP_AR6 dict'inden doldur
        NCV/EF : 1.1, 1.2, 3.3   — yeşil sütunlar boşsa IPCC_FUEL_FACTORS'tan doldur
        Elektrik: 2.1             — EF boşsa TEIAS_KNOWN_FACTORS'tan doldur
        Taşıma : 3.1, 3.2, 3.4   — EF boşsa DEFRA DB'den doldur

    DB'ye sadece DEFRA kaydedilir (online çekilen gerçek veri).
    Diğer tüm değerler Excel'e yazılır → Importer Excel'den okur.

Dolu hücrelere DOKUNULMAZ — kullanıcının manuel değerleri korunur.
"""

import sys
import os
import re
import io
import json
import argparse
import requests
import pandas as pd
from datetime import datetime
from bs4 import BeautifulSoup

sys.path.insert(0, os.path.join(os.path.dirname(__file__)))
from db.connection import SessionLocal
from db.models import EmissionFactor, AuditLog

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; UNSPED-CarbonBot/1.0)"}
TIMEOUT = 30


# ══════════════════════════════════════════════════════════════════
# REFERANS VERİLER
# Bunlar Excel'e yazmak için kullanılır — DB'ye kaydedilmez.
# Sadece DEFRA DB'ye kaydedilir (online çekilen gerçek veri).
# ══════════════════════════════════════════════════════════════════

# IPCC AR6 GWP-100 (Tablo 7.SM.7) — 1.4_Refrigerants GWP sütunu için
GWP_AR6 = {
    "R134A":1526, "R-134A":1526, "HFC134A":1526,
    "R32":771,    "R-32":771,    "HFC32":771,
    "R410A":2088, "R-410A":2088,
    "R407C":1774, "R407c":1774,  "R-407C":1774,
    "R404A":3922, "R-404A":3922,
    "R507A":3985, "R-507A":3985,
    "R22":1960,   "R-22":1960,   "HCFC22":1960,
    "R290":0,     "R600A":4,     "R600":4,
    "R744":1,     "R717":0,
    "FM200":3220, "HFC227EA":3220, "HFC227ea":3220,
    "HFC236FA":8060, "HFC236fa":8060,
    "SF6":25200,
    "R125":3740,  "R-125":3740,
    "R143A":5810, "R-143A":5810,
    "R152A":164,  "R-152A":164,
    "R1234YF":1,  "R1234ZE":1,
    "FK-5-1-12":1, "FK5112":1, "NOVEC1230":1, "C6F12O":1,
    "R448A":1387, "R449A":1397, "R452A":2140,
    "R454B":466,  "R513A":631,
}

# IPCC 2006 GL Vol.2 Table 2.2 — 1.1, 1.2, 3.3 yeşil sütunlar için
IPCC_FUEL_FACTORS = {
    "natural_gas":    {"ncv":48.0,  "co2_factor":56.1, "ch4_factor":0.005,
                       "n2o_factor":0.0001, "density":0.000717,
                       "density_unit":"ton/m³", "input_unit":"m³"},
    "diesel":         {"ncv":43.0,  "co2_factor":74.1, "ch4_factor":0.01,
                       "n2o_factor":0.0006,  "density":0.000832,
                       "density_unit":"ton/L", "input_unit":"L"},
    "petrol":         {"ncv":44.3,  "co2_factor":69.3, "ch4_factor":0.01,
                       "n2o_factor":0.0006,  "density":0.000745,
                       "density_unit":"ton/L", "input_unit":"L"},
    "diesel_shuttle": {"ncv":43.0,  "co2_factor":74.1, "ch4_factor":0.01,
                       "n2o_factor":6e-05,   "density":0.000832,
                       "density_unit":"ton/L", "input_unit":"L"},
}
IPCC_GWP = {"gwp_ch4": 29.8, "gwp_n2o": 273.0}
IPCC_SOURCE = "IPCC 2006 GL Vol.2 Table 2.2"
IPCC_AR = "AR6"

# TC ETKB EVÇED — 2.1_Electricity EF sütunu için
# Kaynak: https://enerji.gov.tr/evced-cevre-ve-iklim-iklim-degisikligi-emisyon-faktorleri
TEIAS_KNOWN_FACTORS = {
    2019: 0.481,
    2020: 0.452,
    2021: 0.434,
    2022: 0.442,
    2023: 0.434,
    2024: 0.434,
}
TEIAS_SOURCE = "TC ETKB EVCED"
TEIAS_URL    = "https://enerji.gov.tr/evced-cevre-ve-iklim-iklim-degisikligi-emisyon-faktorleri"

# DEFRA yedek — online erişilemezse kullanılır
DEFRA_FALLBACK = {
    "All HGVs":        0.08223,
    "HGV refrigerated":0.10453,
    "Average HGV":     0.08223,
    "Small car":       0.14844,
    "Medium car":      0.18432,
    "Large car":       0.27793,
    "Average car":     0.17068,
    "Average van":     0.23339,
    "Domestic flight": 0.25504,
    "Short-haul flight":0.15353,
    "Long-haul flight":0.19085,
}


# ══════════════════════════════════════════════════════════════════
# YARDIMCI FONKSİYONLAR
# ══════════════════════════════════════════════════════════════════

def get_last_check(session, source_name):
    log = session.query(AuditLog).filter(
        AuditLog.action == f"check_{source_name}"
    ).order_by(AuditLog.id.desc()).first()
    if log:
        try: return json.loads(log.notes or "{}")
        except: return {}
    return {}

def save_check(session, source_name, data: dict, status="success"):
    session.add(AuditLog(
        action=f"check_{source_name}",
        scope="emission_factors",
        status=status,
        notes=json.dumps(data, ensure_ascii=False)
    ))

def lookup_gwp(gas_name: str):
    """Gaz ismini birden fazla formatta arar."""
    if not gas_name:
        return None
    attempts = [
        gas_name,
        gas_name.upper(),
        gas_name.upper().replace("-","").replace(" ",""),
        gas_name.strip(),
    ]
    for a in attempts:
        if a in GWP_AR6:
            return GWP_AR6[a]
    return None


# ══════════════════════════════════════════════════════════════════
# KATMAN 1A — IPCC KONTROL (sadece uyarı, değer değişmez)
# ══════════════════════════════════════════════════════════════════

def check_ipcc(session):
    """
    IPCC AR7 yayınlandı mı kontrol eder.
    Yayınlandıysa uyarır — değerleri kendisi güncellemez.
    GWP değerleri IPCC raporuna bağlı olduğundan otomatik güncellenemez,
    geliştirici GWP_AR6 dict'ini manuel güncellemelidir.
    """
    print("\n  Kontrol: IPCC")
    try:
        resp = requests.get(
            "https://www.ipcc.ch/report/ar7/wg1/",
            headers=HEADERS, timeout=TIMEOUT
        )
        if resp.status_code == 200:
            text = resp.text.lower()
            has_real_content = (
                len(resp.text) > 50000 and
                any(k in text for k in ["chapter 7","table 7","global warming potential"])
                and "download" in text
            )
            if has_real_content:
                print("  UYARI: IPCC AR7 WG1 yayinlandi!")
                print("  Yapmaniz gereken: GWP_AR6 sozlugunu guncelle")
                print("  Kaynak: https://www.ipcc.ch/report/ar7/wg1/")
                save_check(session, "ipcc", {"ar7_detected": True,
                           "ts": datetime.now().isoformat()}, status="warning")
                session.commit()
                return
    except Exception as e:
        print(f"  IPCC kontrol hatasi: {e}")

    print(f"  OK: IPCC AR7 henuz yayinlanmadi — {IPCC_AR} gecerli (2029 bekleniyor)")
    save_check(session, "ipcc", {"ar": IPCC_AR,
               "ts": datetime.now().isoformat()})
    session.commit()


# ══════════════════════════════════════════════════════════════════
# KATMAN 1B — TEİAŞ KONTROL (sadece uyarı, değer değişmez)
# ══════════════════════════════════════════════════════════════════

def check_teias(session):
    """
    TEİAŞ/EVÇED sayfasında yeni yıl faktörü var mı kontrol eder.
    Yeni yıl tespit edilirse uyarır — değeri kendisi eklemez.
    Kullanıcı TEIAS_KNOWN_FACTORS'a yeni yılı manuel ekler.
    """
    print("\n  Kontrol: TEIAS/EVCED")
    current_year = datetime.now().year
    last_known   = max(TEIAS_KNOWN_FACTORS.keys())

    if current_year <= last_known:
        print(f"  OK: {last_known} faktoru biliniyor — yeni yil yok")
        return

    # Yeni yıl var — sayfayı kontrol et
    try:
        resp = requests.get(TEIAS_URL, headers=HEADERS, timeout=TIMEOUT)
        if resp.status_code == 200:
            text = resp.text.lower()
            if str(current_year) in text and ("emisyon" in text or "co2" in text):
                print(f"  UYARI: TEIAS {current_year} verisi bulundu!")
                print(f"  Yapmaniz gereken: TEIAS_KNOWN_FACTORS'a {current_year} degerini ekle")
                print(f"  Kaynak: {TEIAS_URL}")
                save_check(session, "teias",
                           {"new_year_detected": current_year,
                            "ts": datetime.now().isoformat()}, status="warning")
                session.commit()
                return
    except Exception as e:
        print(f"  TEIAS erisim hatasi: {e}")

    print(f"  OK: {current_year} icin TEIAS verisi henuz yayinlanmadi")
    save_check(session, "teias", {"last_known": last_known,
               "ts": datetime.now().isoformat()})
    session.commit()


# ══════════════════════════════════════════════════════════════════
# KATMAN 1C — DEFRA FETCHER (online indir + DB'ye kaydet)
# ══════════════════════════════════════════════════════════════════

DEFRA_COLLECTION_URL = (
    "https://www.gov.uk/government/collections/"
    "government-conversion-factors-for-company-reporting"
)

def fetch_defra(session):
    """
    DEFRA'nın en son yılını GOV.UK'dan çeker.
    Yeni yayın varsa indirir, parse eder ve DB'ye kaydeder.
    DB'ye kaydedilen bu değerler Excel'e yazılırken kullanılır.
    """
    print("\n  Kontrol: DEFRA")
    last      = get_last_check(session, "defra")
    last_year = last.get("year", 0)

    # En son yılı bul
    try:
        print("    GOV.UK taranıyor...")
        resp = requests.get(DEFRA_COLLECTION_URL, headers=HEADERS, timeout=TIMEOUT)
        resp.raise_for_status()
        soup  = BeautifulSoup(resp.text, "html.parser")
        pat   = re.compile(
            r"/government/publications/greenhouse-gas-reporting-conversion-factors-(\d{4})"
        )
        years_found = {}
        for link in soup.find_all("a", href=True):
            m = pat.search(link["href"])
            if m:
                y = int(m.group(1))
                years_found[y] = "https://www.gov.uk" + m.group(0)

        if not years_found:
            raise ValueError("Yil sayfasi bulunamadi")

        latest_year = max(years_found.keys())
        latest_url  = years_found[latest_year]
        print(f"    En son DEFRA yili: {latest_year}")

    except Exception as e:
        print(f"    DEFRA erisim hatasi: {e}")
        print("    Yerlesik degerler kullaniliyor...")
        return _save_defra_to_db(session, DEFRA_FALLBACK, datetime.now().year, fallback=True)

    # Zaten güncel mi?
    db_count = session.query(EmissionFactor).filter_by(
        scope="scope3", category="3.x"
    ).count()
    if latest_year == last_year and db_count > 0:
        print(f"    OK: DEFRA {latest_year} guncel — yeni yayin yok ({db_count} kayit)")
        save_check(session, "defra", {"year": latest_year, "status": "no_change"})
        session.commit()
        return 0, 0

    # Flat file'ı indir
    try:
        # Yıl sayfasından flat file URL'sini bul
        yr_resp = requests.get(latest_url, headers=HEADERS, timeout=TIMEOUT)
        yr_resp.raise_for_status()
        yr_soup = BeautifulSoup(yr_resp.text, "html.parser")

        flat_url = None
        for link in yr_soup.find_all("a", href=True):
            href = link["href"]
            text = link.get_text(strip=True).lower()
            if "flat" in text and ("xlsx" in href or "csv" in href):
                flat_url = href if href.startswith("http") else "https://www.gov.uk" + href
                break
        if not flat_url:
            for link in yr_soup.find_all("a", href=True):
                href = link["href"]
                if "assets.publishing.service.gov.uk" in href and "flat" in href.lower():
                    flat_url = href
                    break

        if not flat_url:
            raise ValueError("Flat file URL bulunamadi")

        print(f"    İndiriliyor: {flat_url}")
        dl = requests.get(flat_url, headers=HEADERS, timeout=60)
        dl.raise_for_status()
        print(f"    İndirildi ({len(dl.content)//1024} KB)")

        factors = _parse_defra_flat(dl.content, latest_year)

    except Exception as e:
        print(f"    Flat file hatasi: {e}")
        print("    Yerlesik degerler kullaniliyor...")
        factors = DEFRA_FALLBACK

    return _save_defra_to_db(session, factors, latest_year)

def _parse_defra_flat(content, year):
    """
    DEFRA flat file XLSX parse eder.
    Format: Row 5 = baslik, GHG/Unit='kg CO2e', UOM='km' olan satirlar.
    """
    xl  = pd.ExcelFile(io.BytesIO(content))

    # "Factors by Category" sayfasını bul
    main_sheet = next(
        (n for n in xl.sheet_names
         if any(k in n.lower() for k in ["factor","category","data","ghg"])),
        xl.sheet_names[0]
    )

    df = xl.parse(main_sheet, header=5)
    df.columns = [str(c).strip() for c in df.columns]

    col_level2   = next((c for c in df.columns if "level 2" in c.lower()), None)
    col_level3   = next((c for c in df.columns if "level 3" in c.lower()), None)
    col_col_text = next((c for c in df.columns if "column text" in c.lower()), None)
    col_uom      = next((c for c in df.columns if c.lower() == "uom"), None)
    col_ghg_unit = next((c for c in df.columns if "ghg/unit" in c.lower()), None)
    col_factor   = next((c for c in df.columns if "ghg conversion factor" in c.lower()), None)

    if not col_factor or not col_uom or not col_ghg_unit:
        print("    Sutun bulunamadi — yerlesik degerler kullanilacak")
        return DEFRA_FALLBACK

    # Filtrele: kg CO2e + km
    mask = (
        (df[col_ghg_unit].astype(str).str.strip() == "kg CO2e") &
        (df[col_uom].astype(str).str.strip() == "km")
    )
    df_km = df[mask].copy()
    print(f"    {len(df_km)} km+CO2e satiri bulundu")

    factors = {}
    for _, row in df_km.iterrows():
        level2 = str(row.get(col_level2, "")).strip() if col_level2 else ""
        level3 = str(row.get(col_level3, "nan")).strip() if col_level3 else "nan"
        col_t  = str(row.get(col_col_text,"nan")).strip() if col_col_text else "nan"

        key = f"{level2} — {level3}" if level3 and level3 != "nan" else level2
        if col_t and col_t not in ("nan",""):
            key = f"{key} ({col_t})"

        try:
            val = float(pd.to_numeric(row.get(col_factor), errors="coerce"))
            if 0.0001 < val < 10.0:  # gerçekçi EF aralığı
                factors[key] = val
        except (ValueError, TypeError):
            pass

    if not factors:
        print("    Parse basarisiz — yerlesik degerler kullaniliyor")
        return DEFRA_FALLBACK

    print(f"    {len(factors)} tasimacilik faktoru parse edildi")
    return factors

def _save_defra_to_db(session, factors, year, fallback=False):
    """Parse edilen DEFRA faktörlerini DB'ye kaydeder."""
    source = f"DEFRA {year} GHG Conversion Factors" + (" (fallback)" if fallback else "")
    added = updated = 0
    for vehicle_type, ef in factors.items():
        existing = session.query(EmissionFactor).filter_by(
            fuel_type=vehicle_type, scope="scope3", category="3.x"
        ).first()
        if not existing:
            session.add(EmissionFactor(
                source=source, scope="scope3", category="3.x",
                fuel_type=vehicle_type, region="global",
                unit="kg CO2e/tonne.km", co2_factor=ef,
                input_unit="km", is_active=True, version=1, valid_year=year
            ))
            added += 1
        else:
            if abs((existing.co2_factor or 0) - ef) > 1e-8:
                existing.co2_factor = ef
                existing.valid_year = year
                existing.source = source
                updated += 1

    save_check(session, "defra",
               {"year": year, "added": added, "updated": updated,
                "fallback": fallback, "ts": datetime.now().isoformat()})
    session.commit()
    print(f"  OK: DEFRA — {added} yeni, {updated} guncellendi")
    return added, updated


# ══════════════════════════════════════════════════════════════════
# KATMAN 2 — EXCEL BOŞ HÜCRE DOLDURMA
# ══════════════════════════════════════════════════════════════════

def is_empty(val):
    return val in (None, 0, "", "0", 0.0)

def is_data_row(val):
    if val is None: return False
    s = str(val).strip()
    return not (s.startswith("•") or s.startswith("←") or
                s.startswith("NOTLAR") or len(s) == 0)

def fill_green(cell, value):
    """Boşsa yeşil stilte doldurur. Doluysa dokunmaz. True = değişti."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    if not is_empty(cell.value):
        return False
    thin = Side(style="thin", color="BFBFBF")
    cell.value     = value
    cell.font      = Font(name="Arial", size=10, color="375623")
    cell.fill      = PatternFill("solid", start_color="E2EFDA")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    return True

def fill_excel(excel_path, session):
    """
    Excel'deki boş yeşil hücreleri referans değerlerle doldurur.
    Dolu hücrelere dokunmaz.
    """
    from openpyxl import load_workbook
    print(f"\n  Excel guncelleniyor: {excel_path}")
    wb      = load_workbook(excel_path)
    changes = 0

    # ── 1.4 GWP — IPCC AR6 ───────────────────────────────────────
    ws = wb["1.4_Refrigerants"]
    for row in ws.iter_rows(min_row=4, values_only=False):
        gas_cell = row[1]   # B = gaz türü
        gwp_cell = row[4]   # E = GWP
        if not gas_cell.value or not is_data_row(gas_cell.value):
            continue
        gwp = lookup_gwp(str(gas_cell.value).strip())
        if gwp is not None and fill_green(gwp_cell, gwp):
            print(f"    1.4 {gas_cell.value}: GWP={gwp}")
            changes += 1
        elif gwp is None and is_empty(gwp_cell.value):
            print(f"    UYARI 1.4 {gas_cell.value}: GWP bulunamadi — lutfen manuel girin")

    # ── 1.1, 1.2, 3.3 — IPCC Yakıt Faktörleri ────────────────────
    # (fuel_col, ncv_col, co2_col, ch4_col, n2o_col, gch4_col, gn2o_col, dens_col)
    combustion_sheets = {
        "1.1_Stationary": (2, 5, 6, 7, 8, 9, 10, 11),
        "1.2_Mobile":     (2, 5, 6, 7, 8, 9, 10, 11),
        "3.3_Commuting":  (3, 6, 7, 8, 9, 10, 11, 12),
    }
    for sheet_name, cols in combustion_sheets.items():
        ws = wb[sheet_name]
        fc, ncv_c, co2_c, ch4_c, n2o_c, gch4_c, gn2o_c, den_c = cols
        for row in ws.iter_rows(min_row=4, values_only=False):
            fuel_cell = row[fc - 1]
            if not fuel_cell.value or not is_data_row(fuel_cell.value):
                continue
            fuel = str(fuel_cell.value).strip().lower()
            vals = IPCC_FUEL_FACTORS.get(fuel)
            if not vals:
                continue
            mapping = {
                ncv_c:  vals["ncv"],
                co2_c:  vals["co2_factor"],
                ch4_c:  vals["ch4_factor"],
                n2o_c:  vals["n2o_factor"],
                gch4_c: IPCC_GWP["gwp_ch4"],
                gn2o_c: IPCC_GWP["gwp_n2o"],
                den_c:  vals["density"],
            }
            filled = sum(
                1 for col_idx, val in mapping.items()
                if fill_green(row[col_idx - 1], val)
            )
            if filled:
                print(f"    {sheet_name} [{fuel}]: {filled} hucre dolduruldu")
                changes += filled

    # ── 2.1 Elektrik — TC ETKB EVÇED ─────────────────────────────
    ws = wb["2.1_Electricity"]
    for row in ws.iter_rows(min_row=4, values_only=False):
        year_cell = row[0]
        ef_cell   = row[3]   # D = EF (C artık Birim sütunu)
        if not year_cell.value or not is_data_row(year_cell.value):
            continue
        try:
            year = int(float(str(year_cell.value)))
        except:
            continue

        # Önce DB'den bak (DEFRA değil, TEIAS için de tutarlı olsun)
        ef_rec = session.query(EmissionFactor).filter_by(
            fuel_type="electricity", scope="scope2",
            category="2.1", valid_year=year
        ).first()
        ef = ef_rec.co2_factor if ef_rec else TEIAS_KNOWN_FACTORS.get(year)

        if ef and fill_green(ef_cell, ef):
            print(f"    2.1 {year}: EF={ef} ({TEIAS_SOURCE})")
            changes += 1
        elif not ef and is_empty(ef_cell.value):
            print(f"    UYARI 2.1 {year}: EF bulunamadi — TEIAS_KNOWN_FACTORS'a ekleyin")

    # ── 3.1, 3.2, 3.4 Taşımacılık — DEFRA (DB'den) ───────────────
    defra_records = session.query(EmissionFactor).filter_by(
        scope="scope3", category="3.x"
    ).all()

    # Excel araç ismi → DEFRA DB kaydı eşleştirme
    VEHICLE_MAP = {
        "all hgvs":             ["average hgv", "100% laden"],
        "karayolu-%100 laden":  ["average hgv", "100% laden"],
        "hgv":                  ["average hgv", "100% laden"],
        "small car":            ["small car"],
        "medium car":           ["medium car"],
        "large car":            ["large car"],
        "average car":          ["average car"],
        "average van":          ["average van"],
        "karayolu (dizel)":     ["average car"],
        "karayolu (benzin)":    ["average car"],
        "domestic flight":      ["domestic", "flight"],
        "short-haul flight":    ["short-haul", "flight"],
        "long-haul flight":     ["long-haul", "flight"],
    }

    def find_defra_ef(vehicle_name):
        vl = vehicle_name.lower().strip()
        # 1. Direkt
        for r in defra_records:
            if r.fuel_type.lower() == vl:
                return r.co2_factor
        # 2. Map
        terms = VEHICLE_MAP.get(vl)
        if terms:
            for r in defra_records:
                if all(t in r.fuel_type.lower() for t in terms):
                    return r.co2_factor
        # 3. Kısmi
        words = [w for w in vl.split() if len(w) > 3]
        best, best_score = None, 0
        for r in defra_records:
            score = sum(1 for w in words if w in r.fuel_type.lower())
            if score > best_score:
                best_score = score
                best = r
        if best and best_score >= 1:
            return best.co2_factor
        return None

    transport_sheets = {
        "3.1_Freight":         (5, 8),   # col 7 artık Birim, EF col 8
        "3.2_ProductShipment": (5, 8),
        "3.4_BusinessTravel":  (5, 8),
    }
    for sheet_name, (vcol, ecol) in transport_sheets.items():
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=4, values_only=False):
            v_cell = row[vcol - 1]
            e_cell = row[ecol - 1]
            if not v_cell.value or not is_data_row(v_cell.value):
                continue
            vehicle = str(v_cell.value).strip()
            ef = find_defra_ef(vehicle)
            if ef and is_empty(e_cell.value):
                fill_green(e_cell, ef)
                print(f"    {sheet_name} [{vehicle}]: EF={ef:.5f}")
                changes += 1
            elif ef and not is_empty(e_cell.value):
                print(f"    {sheet_name} [{vehicle}]: {e_cell.value} (dolu, dokunulmuyor)")
            elif not ef:
                print(f"    UYARI {sheet_name} [{vehicle}]: DEFRA eslesme bulunamadi")

    wb.save(excel_path)
    print(f"\n  Excel tamamlandi: {changes} hucre dolduruldu")
    return changes


# ══════════════════════════════════════════════════════════════════
# ANA FONKSİYON
# ══════════════════════════════════════════════════════════════════

def run_update(sources=None, excel_path=None):
    if sources is None:
        sources = ["ipcc", "teias", "defra"]

    print(f"\n{'='*55}")
    print(f"  UNSPED KARBON — FAKTOR GUNCELLEME")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Kaynaklar: {', '.join(s.upper() for s in sources)}")
    print(f"{'='*55}")

    total_added = total_updated = 0

    with SessionLocal() as session:
        # Katman 1: Online kontroller
        if "ipcc" in sources:
            check_ipcc(session)           # sadece uyarır, DB'ye yazmaz

        if "teias" in sources:
            check_teias(session)          # sadece uyarır, DB'ye yazmaz

        if "defra" in sources:
            a, u = fetch_defra(session)   # indirir + DB'ye yazar
            total_added   += a
            total_updated += u

        session.add(AuditLog(
            action="run_update",
            scope="emission_factors",
            status="success",
            notes=json.dumps({
                "sources": sources,
                "db_added":   total_added,
                "db_updated": total_updated,
                "ts": datetime.now().isoformat()
            }, ensure_ascii=False)
        ))
        session.commit()

    # Katman 2: Excel boş hücre doldurma
    excel_changes = 0
    if excel_path:
        if os.path.exists(excel_path):
            with SessionLocal() as session:
                excel_changes = fill_excel(excel_path, session)
        else:
            print(f"\n  UYARI: Excel bulunamadi: {excel_path}")

    print(f"\n{'='*55}")
    print(f"  TAMAMLANDI")
    print(f"  DB (DEFRA)  : {total_added} yeni, {total_updated} guncellendi")
    if excel_path:
        print(f"  Excel       : {excel_changes} hucre dolduruldu")
    print(f"\n  Sonraki adim:")
    print(f"  python pipeline/Importer.py data/UNSPED_Karbon_Veri_Girisi_v3.xlsx")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="UNSPED Karbon — Emisyon faktorlerini guncelle"
    )
    parser.add_argument("--source", "-s", nargs="+",
                        choices=["ipcc","teias","defra","all"],
                        default=["all"],
                        help="Kontrol edilecek kaynaklar")
    parser.add_argument("--excel", "-e", type=str, default=None,
                        help="Excel dosyasi — bos yesil hucreler doldurulur")
    args    = parser.parse_args()
    sources = ["ipcc","teias","defra"] if "all" in args.source else args.source
    run_update(sources=sources, excel_path=args.excel)